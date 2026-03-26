"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   SCRIPT_TOP_Universal_Caracterizacion_Excel.py                            ║
║   Genera tablas de caracterización al ingreso (TOP1)                       ║
║   Compatible con cualquier país que use el instrumento TOP                 ║
║   Versión Universal 1.0                                                    ║
╠══════════════════════════════════════════════════════════════════════════════╣
║                                                                             ║
║  CÓMO USAR LA PRÓXIMA VEZ:                                                 ║
║  1. Abre un chat nuevo con Claude                                           ║
║  2. Sube DOS archivos:                                                      ║
║       • Este script: SCRIPT_TOP_Universal_Caracterizacion_Excel.py         ║
║       • La base en formato Wide (generada por SCRIPT_TOP_Universal_Wide)   ║
║  3. Escribe exactamente:                                                    ║
║     "Ejecuta el script universal Caracterización Excel con esta base Wide" ║
║                                                                             ║
║  TABLAS GENERADAS (11 tablas, solo TOP1 — ingreso):                        ║
║    1.1  Distribución por Sexo                                               ║
║    1.2  Distribución por Rango de Edad                                      ║
║    2.1  Consumo Sustancia Principal                                         ║
║    2.2  Promedio Días Consumo – Sustancia Principal                         ║
║    2.3  Consumo de Sustancias (% personas, puede >100%)                    ║
║    2.4  Promedio Días de Consumo por Sustancia                              ║
║    3.1  Transgresión a la Norma Social                                      ║
║    3.2  Distribución por Tipo de Transgresión                               ║
║    4.1  Autopercepción del Estado de Salud (0–20)                           ║
║    4.2  Condiciones de Vivienda                                             ║
║    5.   Días Trabajados y Días de Estudio                                   ║
║                                                                             ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""


import glob, os, unicodedata

def _norm(s):
    return unicodedata.normalize('NFD', str(s).lower()).encode('ascii','ignore').decode()

def auto_archivo_wide():
    """Encuentra automáticamente la base Wide subida al chat"""
    candidatos = (
        glob.glob('/mnt/user-data/uploads/*Wide*.xlsx') +
        glob.glob('/mnt/user-data/uploads/*wide*.xlsx') +
        glob.glob('/mnt/user-data/uploads/TOP_Base*.xlsx') +
        glob.glob('/home/claude/TOP_Base_Wide.xlsx'))
    if not candidatos:
        raise FileNotFoundError(
            "\n\u26a0  No se encontró la base Wide.\n"
            "   Sube el archivo TOP_Base_Wide.xlsx junto con este script.")
    print(f"  \u2192 Base Wide detectada: {os.path.basename(candidatos[0])}")
    return candidatos[0]

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN — Claude ajusta NOMBRE_SERVICIO y PERIODO según corresponda
# ══════════════════════════════════════════════════════════════════════════════

SHEET_NAME  = 'Base Wide'
OUTPUT_FILE = '/home/claude/TOP_Caracterizacion_Ingreso.xlsx'
INPUT_FILE  = auto_archivo_wide()
# ── FILTRO OPCIONAL POR CENTRO ────────────────────────────────────────────────
# Dejar en None para procesar TODOS los centros.
# Poner el código exacto del centro para filtrar solo ese centro.
# Ejemplos:
#   FILTRO_CENTRO = None         ← todos los centros
#   FILTRO_CENTRO = "HCHN01"     ← solo ese centro
FILTRO_CENTRO = None
# ─────────────────────────────────────────────────────────────────────────────
   # ← detecta automáticamente la base Wide

# ══════════════════════════════════════════════════════════════════════════════
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import warnings
warnings.filterwarnings('ignore')

def _es_positivo(valor):
    s = str(valor).strip().lower()
    if s in ('sí', 'si'): return True
    if s in ('no', 'no aplica', 'nunca', 'nan', ''): return False
    n = pd.to_numeric(valor, errors='coerce')
    return not pd.isna(n) and n > 0

C_DARK='1F3864'; C_MID='2E75B6'; C_LIGHT='BDD7EE'; C_WHITE='FFFFFF'
C_ALT='EEF4FB';  C_NOTE='595959'; C_BDR='B8CCE4'

thin = Side(style='thin', color=C_BDR)
def bd(): return Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Helpers de formato ────────────────────────────────────────────────────────
def sec(ws, row, num, title):
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f'B{row}:F{row}')
    c = ws[f'B{row}']; c.value = f'{num}.  {title}'
    c.font = Font(bold=True, size=11, color=C_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color=C_MID)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    for col in range(2, 7): ws.cell(row, col).border = bd()
    return row + 1

def hdrs(ws, row, labels):
    ws.row_dimensions[row].height = 20
    for cl, lbl in zip(['B','C','D','E','F'], labels):
        c = ws[f'{cl}{row}']; c.value = lbl
        c.font = Font(bold=True, size=9, color=C_DARK, name='Arial')
        c.fill = PatternFill('solid', start_color=C_LIGHT)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bd()
    return row + 1

def drow(ws, row, vals, alt=False, ba=False, dark=False):
    ws.row_dimensions[row].height = 16
    bg = C_DARK if dark else (C_ALT if alt else C_WHITE)
    tc = C_WHITE if dark else '000000'
    for i, (cl, val) in enumerate(zip(['B','C','D','E','F'], vals)):
        c = ws[f'{cl}{row}']
        c.value = round(val, 1) if isinstance(val, float) and not np.isnan(val) else val
        c.font = Font(size=9, name='Arial', color=tc, bold=(ba or dark))
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='left' if i==0 else 'center',
                                 vertical='center', indent=1 if i==0 else 0)
        c.border = bd()
    return row + 1

def note(ws, row, text):
    ws.merge_cells(f'B{row}:F{row}')
    c = ws[f'B{row}']; c.value = text
    c.font = Font(size=8, color=C_NOTE, name='Arial', italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 13
    return row + 2

def safe_mean(s):
    c = pd.to_numeric(s, errors='coerce')
    return c.mean(), int(c.notna().sum())

# ── Normalización sustancia principal ─────────────────────────────────────────
def norm_sust(s):
    if pd.isna(s) or str(s).strip() == '0': return None
    s = str(s).strip().lower()
    if any(x in s for x in ['alcohol','cerveza','licor','aguard','alchol','bebida']): return 'Alcohol'
    if any(x in s for x in ['marihu','marjhu','marhuana','cannabis','cannbis','cannabin']): return 'Cannabis/Marihuana'
    if any(x in s for x in ['crack','cristal','piedra','paco']): return 'Crack/Cristal'
    if any(x in s for x in ['pasta base','pasta','papelillo']): return 'Pasta Base'
    if any(x in s for x in ['cocain','perico','coca ']): return 'Cocaína'
    if any(x in s for x in ['tabaco','cigarr','nicot']): return 'Tabaco/Nicotina'
    if any(x in s for x in ['inhalant','thiner','activo','resistol','cemento']): return 'Inhalantes'
    if any(x in s for x in ['sedant','benzod','tranqui','valium','clonaz']): return 'Sedantes'
    if any(x in s for x in ['opiod','heroina','morfin','fentanil']): return 'Opiáceos'
    if any(x in s for x in ['metanfet','anfetam']): return 'Metanfetamina'
    return 'Otras'

# ══════════════════════════════════════════════════════════════════════════════
# DETECCIÓN DINÁMICA DE COLUMNAS (trabaja sobre columnas _TOP1)
# ══════════════════════════════════════════════════════════════════════════════
def detectar_columnas(cols):
    """Detecta automáticamente columnas clave en la base Wide (_TOP1)."""

    def t1(c): return f'{c}_TOP1'  # helper para agregar sufijo

    # 1. Sustancias: columnas "1) Registrar..." + "Total (0-28)_TOP1"
    sust_cols = []
    for c in cols:
        if c.endswith('_TOP1') and 'Total (0-28)' in c:
            base = c.replace('_TOP1','')
            if base.startswith('1)'):
                partes = base.split('>>')
                if len(partes) >= 3:
                    nombre = partes[-2].strip().split('(')[0].strip()
                    sust_cols.append((nombre, c))
    print(f'  Sustancias detectadas: {[s[0] for s in sust_cols]}')

    # 2. Transgresión Sí/No
    tr_sn = []
    for c in cols:
        if c.endswith('_TOP1') and c.replace('_TOP1','').startswith('3)') and '>>' in c:
            nombre = c.replace('_TOP1','').split('>>')[-1].strip()
            tr_sn.append((nombre, c))
    print(f'  Transgresión tipos: {[t[0] for t in tr_sn]}')

    # 3. VIF
    vif = next((c for c in cols if c.endswith('_TOP1') and '4)' in c
                and 'Violencia Intrafamiliar' in c and 'Total (0-28)' in c), None)

    # 4. Salud
    sal_psi = next((c for c in cols if c.endswith('_TOP1') and c.replace('_TOP1','').startswith('6)')), None)
    sal_fis = next((c for c in cols if c.endswith('_TOP1') and c.replace('_TOP1','').startswith('8)')), None)
    cal_vid = next((c for c in cols if c.endswith('_TOP1') and c.replace('_TOP1','').startswith('10)')), None)

    # 5. Vivienda
    viv1 = next((c for c in cols if c.endswith('_TOP1') and '9)' in c and 'estable' in c.lower()), None)
    viv2 = next((c for c in cols if c.endswith('_TOP1') and '9)' in c and 'básicas' in c.lower()), None)

    # 6. Trabajo y estudio
    trab  = next((c for c in cols if c.endswith('_TOP1') and '7)' in c
                  and 'trabajo remunerado' in c.lower() and 'Total (0-28)' in c), None)
    estud = next((c for c in cols if c.endswith('_TOP1') and '7)' in c
                  and ('colegio' in c.lower() or 'instituto' in c.lower())
                  and 'Total (0-28)' in c), None)

    # 7. Sustancia principal
    sust_ppal = next((c for c in cols if c.endswith('_TOP1')
                      and c.replace('_TOP1','').startswith('2)')
                      and 'sustancia principal' in c.lower()), None)

    # 8. Sexo y fecha de nacimiento
    sexo   = next((c for c in cols if c.endswith('_TOP1') and 'sexo' in c.lower()), None)
    fn     = next((c for c in cols if c.endswith('_TOP1') and 'nacimiento' in c.lower()), None)
    fecha  = next((c for c in cols if c.endswith('_TOP1') and 'fecha entrevista' in c.lower()), None)

    DC = dict(sust_cols=sust_cols, tr_sn=tr_sn, vif=vif,
              sal_psi=sal_psi, sal_fis=sal_fis, cal_vid=cal_vid,
              viv1=viv1, viv2=viv2, trab=trab, estud=estud,
              sust_ppal=sust_ppal, sexo=sexo, fn=fn, fecha=fecha)

    for k, v in DC.items():
        if v is None and k not in ('sust_cols', 'tr_sn'):
            print(f'  ⚠️  No encontrada: {k}')
    return DC

# ══════════════════════════════════════════════════════════════════════════════
# CARGA — lee base Wide, toma columnas _TOP1 (todos los pacientes = ingreso)
# ══════════════════════════════════════════════════════════════════════════════
def cargar_ingreso():
    print(f'  Leyendo: {INPUT_FILE}  |  Hoja: {SHEET_NAME}')
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=1)  # fila 2 = encabezados

    # Aplicar filtro de centro si corresponde
    def _norm(s):
        import unicodedata
        return unicodedata.normalize('NFKD', s.lower()).encode('ascii','ignore').decode()
    _col_centro = next((c for c in df.columns if any(x in _norm(c) for x in
                        ['codigo del centro', 'servicio de tratamiento', 'centro/ servicio',
                         'codigo centro'])), None)
    if FILTRO_CENTRO and _col_centro:
        n_antes = len(df)
        df = df[df[_col_centro].astype(str).str.strip() == FILTRO_CENTRO].copy()
        df = df.reset_index(drop=True)
        print(f'  ⚑ Filtro activo: Centro = "{FILTRO_CENTRO}"')
        print(f'    {n_antes} pacientes totales → {len(df)} del centro seleccionado')
    if FILTRO_CENTRO:
        global OUTPUT_FILE
        OUTPUT_FILE = f'/home/claude/TOP_Caracterizacion_{FILTRO_CENTRO}.xlsx'
    N  = len(df)
    print(f'  Pacientes (ingreso TOP1): {N}')
    return df, N

# ══════════════════════════════════════════════════════════════════════════════
# CONSTRUCCIÓN DEL EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def build_report(wb, d, N, DC):
    ws = wb.active; ws.title = 'Caracterización Ingreso'
    ws.sheet_properties.tabColor = C_MID
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 46
    for col in ['C','D','E','F']: ws.column_dimensions[col].width = 16

    # Encabezado
    for r, txt, bg, sz, bold in [
        (1, 'CARACTERIZACIÓN DE PACIENTES AL INGRESO  ·  TOP', C_DARK, 16, True),
        (2, f'Primera aplicación del instrumento TOP  ·  N = {N} pacientes', C_MID, 10, False),
        (3, f'Base total: {N} pacientes  ·  % calculados sobre N válido de cada variable', C_LIGHT, 9, False)]:
        ws.merge_cells(f'B{r}:F{r}')
        c = ws[f'B{r}']; c.value = txt
        c.font = Font(bold=bold, size=sz, color=C_WHITE if r<3 else C_DARK, name='Arial')
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 36 if r==1 else 20 if r==2 else 14

    R = 5

    # ── SECCIÓN 1: ANTECEDENTES GENERALES ─────────────────────────────────────
    R = sec(ws, R, '1', 'Antecedentes Generales')

    # 1.1 Sexo
    R = sec(ws, R, '1.1', 'Distribución por Sexo')
    R = hdrs(ws, R, ['Sexo', 'n', '%', '', ''])
    if DC['sexo']:
        sc = d[DC['sexo']].astype(str).str.strip().str.upper()
        nv_sex = int(sc.isin(['H','M']).sum())
        n_h = int((sc=='H').sum()); n_m = int((sc=='M').sum())
        R = drow(ws, R, ['Hombre', n_h, round(n_h/nv_sex*100,1) if nv_sex>0 else 0, '', ''], alt=False)
        R = drow(ws, R, ['Mujer',  n_m, round(n_m/nv_sex*100,1) if nv_sex>0 else 0, '', ''], alt=True)
        R = drow(ws, R, ['N válido', nv_sex, '100%', '', ''], ba=True)
        R = note(ws, R, f'N válido = {nv_sex} personas con sexo registrado (H=Hombre, M=Mujer).')
    else:
        R = note(ws, R, 'Columna Sexo no encontrada en la base.')

    # 1.2 Rango de edad
    R = sec(ws, R, '1.2', 'Distribución por Rango de Edad')
    R = hdrs(ws, R, ['Rango de edad', 'n', '%', 'N válido', ''])
    if DC['fn'] and DC['fecha']:
        fn_s  = pd.to_datetime(d[DC['fn']], errors='coerce')
        ref_s = pd.to_datetime(d[DC['fecha']], errors='coerce').fillna(pd.Timestamp.now())
        edad  = ((ref_s - fn_s).dt.days / 365.25).round(1)
        edad  = edad[(edad >= 10) & (edad <= 100)]
        nv_edad = int(edad.notna().sum())
        bins = [0, 17, 30, 40, 50, 60, 200]
        labs = ['Menos de 18 años','18 a 30 años','31 a 40 años',
                '41 a 50 años','51 a 60 años','61 o más años']
        ec = pd.cut(edad, bins=bins, labels=labs)
        for i, l in enumerate(labs):
            n_l = int((ec==l).sum())
            R = drow(ws, R, [l, n_l, round(n_l/nv_edad*100,1) if nv_edad>0 else 0,
                             f'n={nv_edad}', ''], alt=i%2==0)
        R = drow(ws, R, ['N válido', nv_edad, '100%', '', ''], ba=True)
        prom_e = round(float(edad.mean()),1); de_e = round(float(edad.std()),1)
        emin = int(edad.min()) if edad.notna().any() else 0
        emax = int(edad.max()) if edad.notna().any() else 0
        R = note(ws, R, f'Edad promedio: {prom_e} años (DE={de_e}; rango {emin}–{emax}).')
    else:
        R = note(ws, R, 'Columnas de fecha de nacimiento o fecha de entrevista no encontradas.')

    # ── SECCIÓN 2: CONSUMO DE SUSTANCIAS ──────────────────────────────────────
    R = sec(ws, R, '2', 'Consumo de Sustancias')

    # 2.1 Sustancia principal
    R = sec(ws, R, '2.1', 'Consumo Sustancia Principal')
    R = hdrs(ws, R, ['Sustancia', 'n', '%', 'N válido', ''])
    if DC['sust_ppal']:
        sr   = d[DC['sust_ppal']].apply(norm_sust).dropna()
        nv_s = len(sr); vc = sr.value_counts()
        cats = ['Alcohol','Cannabis/Marihuana','Pasta Base','Cocaína','Crack/Cristal',
                'Tabaco/Nicotina','Inhalantes','Sedantes','Opiáceos','Metanfetamina','Otras']
        for i, cat in enumerate(cats):
            n_c = int(vc.get(cat, 0))
            if n_c == 0: continue
            R = drow(ws, R, [cat, n_c, round(n_c/nv_s*100,1) if nv_s>0 else 0,
                             f'n={nv_s}', ''], alt=i%2==0)
        R = drow(ws, R, ['N válido', nv_s, '100%', '', ''], ba=True)
        R = note(ws, R, f'% sobre casos con sustancia identificada. N válido = {nv_s}')
    else:
        R = note(ws, R, 'Columna sustancia principal no encontrada.')

    # 2.2 Días de consumo — SUSTANCIA PRINCIPAL
    R = sec(ws, R, '2.2', 'Promedio de Días de Consumo – Sustancia Principal (últimas 4 semanas)')
    R = hdrs(ws, R, ['Sustancia', 'Promedio días', 'N (declararon\ncomo principal)', 'N válido', ''])
    if DC['sust_ppal'] and DC['sust_cols']:
        sust_norm_s = d[DC['sust_ppal']].apply(norm_sust)
        for i, (lbl, col) in enumerate(DC['sust_cols']):
            v    = pd.to_numeric(d[col], errors='coerce')
            mask = sust_norm_s.apply(
                lambda s: isinstance(s, str) and lbl.lower() in s.lower().replace('\n',' '))
            sub  = v[mask & (v > 0)]
            if len(sub) == 0: continue
            R = drow(ws, R, [lbl, round(float(sub.mean()),1), int(len(sub)),
                             int(v.notna().sum()), ''], alt=i%2==0)
        R = note(ws, R, 'Promedio calculado entre quienes declararon esa sustancia como su principal problema.')
    else:
        R = note(ws, R, 'Columnas de sustancias no detectadas.')

    # 2.3 Consumo de sustancias (% personas, puede sumar >100%)
    R = sec(ws, R, '2.3', 'Consumo de Sustancias (% de personas que consume cada sustancia)')
    R = hdrs(ws, R, ['Sustancia', 'n consumidores', '% sobre N total', 'N total', ''])
    for i, (lbl, col) in enumerate(DC['sust_cols']):
        v   = pd.to_numeric(d[col], errors='coerce')
        n_c = int((v > 0).sum())
        if n_c == 0: continue
        R = drow(ws, R, [lbl, n_c, round(n_c/N*100,1), N, ''], alt=i%2==0)
    R = note(ws, R, f'% sobre N total = {N} pacientes. Los % pueden sumar más de 100% (una persona puede consumir varias sustancias).')

    # 2.4 Promedio días por sustancia (todos los consumidores)
    R = sec(ws, R, '2.4', 'Promedio de Días de Consumo por Sustancia (últimas 4 semanas)')
    R = hdrs(ws, R, ['Sustancia', 'Promedio días', 'N consumidores\n(días > 0)', 'N válido', ''])
    for i, (lbl, col) in enumerate(DC['sust_cols']):
        v   = pd.to_numeric(d[col], errors='coerce')
        n_c = int((v > 0).sum())
        if n_c == 0: continue
        prom = round(float(v[v>0].mean()), 1)
        R = drow(ws, R, [lbl, prom, n_c, int(v.notna().sum()), ''], alt=i%2==0)
    R = note(ws, R, 'Promedio calculado solo entre consumidores (días > 0), independiente de cuál sea su sustancia principal.')

    # ── SECCIÓN 3: TRANSGRESIÓN ────────────────────────────────────────────────
    R = sec(ws, R, '3', 'Transgresión a la Norma Social')

    tr_cols = [c for _, c in DC['tr_sn']]

    def has_tr(row):
        for c in tr_cols:
            if _es_positivo(row.get(c, '')): return True
        if DC['vif']:
            v = pd.to_numeric(row.get(DC['vif'], np.nan), errors='coerce')
            return not np.isnan(v) and v > 0
        return False

    def valid_tr(row):
        for c in tr_cols:
            if pd.notna(row.get(c)): return True
        if DC['vif']:
            return pd.notna(row.get(DC['vif']))
        return False

    t    = d.apply(lambda r: int(has_tr(r)), axis=1)
    vl   = d.apply(lambda r: int(valid_tr(r)), axis=1)
    nv_t = int(vl.sum()); n_si = int(t.sum()); n_no = nv_t - n_si

    # 3.1 Transgresión general
    R = sec(ws, R, '3.1', 'Transgresión a la Norma Social (presencia de algún incidente)')
    R = hdrs(ws, R, ['Situación', 'n', '%', 'N válido', ''])
    R = drow(ws, R, ['Sí, cometió alguna transgresión', n_si,
                     round(n_si/nv_t*100,1) if nv_t>0 else 0, f'n={nv_t}', ''], alt=False)
    R = drow(ws, R, ['No, ninguna transgresión', n_no,
                     round(n_no/nv_t*100,1) if nv_t>0 else 0, f'n={nv_t}', ''], alt=True)
    R = drow(ws, R, ['N válido', nv_t, '100%', '', ''], ba=True)
    tipos_str = ', '.join(n for n,_ in DC['tr_sn'])
    R = note(ws, R, f'Incluye {tipos_str} y Violencia Intrafamiliar (VIF).')

    # 3.2 Tipos de transgresión
    R = sec(ws, R, '3.2', 'Distribución por Tipo de Transgresión (% sobre total de transgresores)')
    R = hdrs(ws, R, ['Tipo de transgresión', 'n', '% del total\nque transgredió', f'N base\n(transgresores)', ''])
    for i, (lbl, col) in enumerate(DC['tr_sn']):
        n = int(d[col].apply(_es_positivo).sum())
        R = drow(ws, R, [lbl, n, round(n/n_si*100,1) if n_si>0 else 0, n_si, ''], alt=i%2==0)
    if DC['vif']:
        vif_v = pd.to_numeric(d[DC['vif']], errors='coerce')
        n_vif = int((vif_v > 0).sum())
        R = drow(ws, R, ['Violencia Intrafamiliar (VIF)', n_vif,
                         round(n_vif/n_si*100,1) if n_si>0 else 0, n_si, ''],
                 alt=len(DC['tr_sn'])%2==0)
    R = note(ws, R, f'% sobre los {n_si} pacientes que cometieron al menos una transgresión. '
                    'Un mismo paciente puede aparecer en más de un tipo (% no suman 100%).')

    # ── SECCIÓN 4: SALUD, CALIDAD DE VIDA Y VIVIENDA ──────────────────────────
    R = sec(ws, R, '4', 'Salud, Calidad de Vida y Vivienda')

    # 4.1 Salud
    R = sec(ws, R, '4.1', 'Autopercepción del Estado de Salud (escala 0–20)')
    R = hdrs(ws, R, ['Dimensión', 'Promedio', 'Mínimo', 'Máximo', 'N válido'])
    for i, (lbl, col) in enumerate([
        ('Salud Psicológica', DC['sal_psi']),
        ('Salud Física',      DC['sal_fis']),
        ('Calidad de Vida',   DC['cal_vid']),
    ]):
        if col is None: continue
        v = pd.to_numeric(d[col], errors='coerce')
        m, nv = safe_mean(d[col])
        R = drow(ws, R, [lbl, round(m,1), round(float(v.min()),0),
                         round(float(v.max()),0), nv], alt=i%2==0)
    R = note(ws, R, 'Escala 0 = Muy mal / 20 = Excelente. A mayor puntaje, mejor estado percibido.')

    # 4.2 Vivienda
    R = sec(ws, R, '4.2', 'Condiciones de Vivienda al Ingreso (últimas 4 semanas)')
    R = hdrs(ws, R, ['Condición', 'n Sí', '% Sí', 'n No', 'N válido'])
    for i, (lbl, col) in enumerate([
        ('Tiene lugar estable para vivir',  DC['viv1']),
        ('Vivienda con condiciones básicas', DC['viv2']),
    ]):
        if col is None: continue
        nv  = int(d[col].isin(['Sí','No']).sum()) or N
        n_s = int((d[col] == 'Sí').sum()); n_n = int((d[col] == 'No').sum())
        R = drow(ws, R, [lbl, n_s, round(n_s/nv*100,1), n_n, nv], alt=i%2==0)
    R = note(ws, R, '% sobre casos con respuesta Sí/No válida.')

    # ── SECCIÓN 5: TRABAJO Y ESTUDIO ──────────────────────────────────────────
    R = sec(ws, R, '5', 'Días Trabajados y Días de Estudio al Ingreso (últimas 4 semanas, 0–28)')
    R = hdrs(ws, R, ['Actividad', 'Promedio\ndías', 'N con\nactividad (>0)', 'N válido', ''])
    for i, (lbl, col) in enumerate([
        ('Días de trabajo remunerado',       DC['trab']),
        ('Días asistidos a inst. educativa', DC['estud']),
    ]):
        if col is None: continue
        v    = pd.to_numeric(d[col], errors='coerce')
        nv   = int(v.notna().sum()); n_act = int((v > 0).sum()); m = v.mean()
        R = drow(ws, R, [lbl, round(m,1) if not np.isnan(m) else 0, n_act, nv, ''], alt=i%2==0)
    R = note(ws, R, 'Promedio sobre todos los pacientes (incluye quienes no trabajan/estudian = 0 días).')

    ws.freeze_panes = 'B5'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    print(f'  ✓ 11 tablas generadas  ·  N = {N} pacientes')

# ══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    print('=' * 60)
    print('  SCRIPT_TOP_Universal_Caracterizacion_Excel  —  Iniciando...')
    print('=' * 60)

    d, N = cargar_ingreso()

    print('\n→ Detectando columnas dinámicamente...')
    DC = detectar_columnas(d.columns.tolist())

    print('\n→ Construyendo Excel...')
    wb = Workbook()
    build_report(wb, d, N, DC)
    wb.save(OUTPUT_FILE)

    print(f'\n{"=" * 60}')
    print(f'  ✅  LISTO  →  {OUTPUT_FILE}')
    print(f'{"=" * 60}')
