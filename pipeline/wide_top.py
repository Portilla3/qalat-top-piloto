"""
pipeline/wide_top.py
Motor de procesamiento TOP — adaptado de SCRIPT_TOP_Universal_Wide_v3_6.py
Expone una función principal: procesar_wide(input_path) → dict con resultados
"""
import pandas as pd
import numpy as np
import re, unicodedata, warnings
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings('ignore')

# ── Colores Excel ─────────────────────────────────────────────────────────────
C_DARK  = '1F3864'; C_MID = '2E75B6'; C_TOP2 = '00B0F0'
C_LIGHT = 'BDD7EE'; C_ALT = 'EEF4FB'; C_WHITE = 'FFFFFF'; C_BDR = 'B8CCE4'

MESES_ES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
            7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

_MES_ES = {'ene':'Jan','feb':'Feb','mar':'Mar','abr':'Apr','may':'May','jun':'Jun',
           'jul':'Jul','ago':'Aug','sept':'Sep','sep':'Sep','oct':'Oct',
           'nov':'Nov','dic':'Dec'}
_EXCEL_ORIGEN = pd.Timestamp('1899-12-30')

_SUST_KEYS_NORM = [_norm_str(x) for x in [
    'sustancia principal', 'cual considera', 'cuál considera',
    'genera mas problemas', 'genera más problemas'
]] if False else []  # se inicializa abajo

def _norm_str(s):
    return unicodedata.normalize('NFD', str(s).lower()).encode('ascii','ignore').decode()

_SUST_KEYS_NORM = [_norm_str(x) for x in [
    'sustancia principal', 'cual considera', 'cuál considera',
    'genera mas problemas', 'genera más problemas'
]]

def _parse_fecha(serie):
    if pd.api.types.is_datetime64_any_dtype(serie): return serie
    if pd.api.types.is_numeric_dtype(serie):
        return _EXCEL_ORIGEN + pd.to_timedelta(serie.fillna(0).astype(int), unit='D')
    def _conv(val):
        s = str(val).strip().lower()
        for es, en in _MES_ES.items():
            s = re.sub(rf'\b{es}\b', en, s)
        return pd.to_datetime(s, errors='coerce')
    result = pd.to_datetime(serie, errors='coerce')
    mask_nat = result.isna() & serie.notna()
    if mask_nat.any():
        result[mask_nat] = serie[mask_nat].apply(_conv)
    return result

def norm_sust_v3(s):
    if pd.isna(s): return None
    raw = str(s).strip()
    if raw in ('0', ''): return None
    raw = re.split(r'[\r\n]', raw)[0].strip()
    raw = re.sub(r'\(.*?\)', '', raw).strip()
    raw = re.sub(r'^(las dos|ambas|los dos|ambos)[,\s]+', '', raw, flags=re.IGNORECASE).strip()
    primera = re.split(r'\s+y\s+|[/,+]', raw, maxsplit=1)[0].strip()
    n = _norm_str(primera)
    if any(x in n for x in ['ninguno','ninguna','niega','no aplica','no consume','nada']): return None
    if any(x in n for x in ['ludopatia','juego','apuesta','gaming','azar']): return None
    if any(x in n for x in ['alcohol','alchol','cerveza','licor','aguard','beer','wine','ron']): return 'Alcohol'
    if any(x in n for x in ['marihu','marhuana','cannabis','cannbis','marij','weed','crispy']): return 'Marihuana'
    if any(x in n for x in ['tusi','tussi','tusy','tuci','2cb']): return 'Tusi'
    if any(x in n for x in ['pasta base','pasta basica','papelillo','pbc','basuco','bazuco']): return 'Pasta Base/Basuco'
    if any(x in n for x in ['metanfet','anfetam','cristal','crystal']): return 'Metanfetamina'
    if any(x in n for x in ['crack','piedra','paco']): return 'Crack'
    if any(x in n for x in ['cocain','cocai','perico','coke']): return 'Cocaína'
    if any(x in n for x in ['tabaco','cigarr','nicot']): return 'Tabaco'
    if any(x in n for x in ['inhalant','thiner','activo','pegamento','solvente']): return 'Inhalantes'
    if any(x in n for x in ['sedant','benzod','tranqui','clonaz','diazep','rivotril']): return 'Sedantes'
    if any(x in n for x in ['opiod','heroina','morfin','fentanil','tramad']): return 'Opiáceos'
    if any(x in n for x in ['extasis','mdma','xtc']): return 'Éxtasis'
    if any(x in n for x in ['ketam']): return 'Ketamina'
    return None

def auto_col(cols, keywords, nombre_col):
    for c in cols:
        if any(_norm_str(k) in _norm_str(c) for k in keywords):
            return c
    raise ValueError(f"Columna '{nombre_col}' no encontrada. Palabras clave: {keywords}\n"
                     f"Columnas disponibles: {list(cols)[:20]}")

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
def procesar_wide(input_path: str,
                  filtro_centro: str = None,
                  fecha_desde: str = None,
                  fecha_hasta: str = None) -> dict:
    """
    Procesa una base bruta TOP y retorna un diccionario con resultados.

    Args:
        input_path:    ruta al Excel bruto
        filtro_centro: código exacto del centro (None = todos)
        fecha_desde:   'YYYY-MM' inicio del período (None = sin límite)
        fecha_hasta:   'YYYY-MM' fin del período    (None = sin límite)
    """
    logs = []

    # ── Carga ────────────────────────────────────────────────────────────────
    df = pd.read_excel(input_path, sheet_name=0, header=0)
    logs.append(f"✓ Archivo cargado: {len(df)} filas × {len(df.columns)} columnas")

    COL_CODIGO = auto_col(df.columns,
        keywords=['identificacion','identificación','2 primeras letras','primer nombre','cod_pac','id_pac'],
        nombre_col='Código de identificación del paciente')
    COL_FECHA = auto_col(df.columns,
        keywords=['fecha entrevista','fecha_entrevista','fecha de entrevista','fechaentrevista'],
        nombre_col='Fecha de Entrevista')
    logs.append(f"✓ Columna código: {COL_CODIGO[:50]}")
    logs.append(f"✓ Columna fecha: {COL_FECHA}")

    # ── Validación ────────────────────────────────────────────────────────────
    hoy = pd.Timestamp.now()
    df[COL_FECHA] = _parse_fecha(df[COL_FECHA])
    alertas = []

    COL_CENTRO = None
    for c in df.columns:
        nc = _norm_str(c)
        if any(k in nc for k in ['codigo del centro','centro de tratamiento','servicio de tratamiento']):
            if 'trabajo' not in nc and 'estudio' not in nc:
                COL_CENTRO = c; break

    # ── Filtro por CENTRO ─────────────────────────────────────────────────────
    n_antes = len(df)
    if filtro_centro and COL_CENTRO:
        df = df[df[COL_CENTRO].astype(str).str.strip() == filtro_centro.strip()].copy()
        df = df.reset_index(drop=True)
        logs.append(f"✓ Filtro centro '{filtro_centro}': {n_antes} → {len(df)} filas")
        if len(df) == 0:
            raise ValueError(f"El centro '{filtro_centro}' no tiene registros. "
                             f"Verifica el código exacto.")

    # ── Filtro por PERÍODO ────────────────────────────────────────────────────
    if fecha_desde or fecha_hasta:
        mask_periodo = pd.Series([True] * len(df), index=df.index)
        if fecha_desde:
            desde_ts = pd.Timestamp(fecha_desde + '-01')
            mask_periodo &= df[COL_FECHA] >= desde_ts
        if fecha_hasta:
            hasta_ts = pd.Timestamp(fecha_hasta + '-01') + pd.offsets.MonthEnd(0)
            mask_periodo &= df[COL_FECHA] <= hasta_ts
        n_antes2 = len(df)
        df = df[mask_periodo].copy().reset_index(drop=True)
        periodo_label = f"{fecha_desde or '?'} → {fecha_hasta or '?'}"
        logs.append(f"✓ Filtro período {periodo_label}: {n_antes2} → {len(df)} filas")
        if len(df) == 0:
            raise ValueError(f"No hay registros en el período seleccionado ({periodo_label}).")

    centro_lookup = {}
    if COL_CENTRO:
        centro_lookup = df.groupby(COL_CODIGO)[COL_CENTRO].first().to_dict()

    def get_centro(cod):
        return str(centro_lookup.get(cod, '—'))[:60]

    # Fecha nacimiento
    COL_FN = None
    for c in df.columns:
        if c == COL_CODIGO: continue
        nc = _norm_str(c)
        if any(k in nc for k in ['fecha de nacimiento','fecha_nacimiento','fecha nac','fechanac']):
            COL_FN = c; break

    if COL_FN:
        df[COL_FN] = _parse_fecha(df[COL_FN])
        for idx, row in df.iterrows():
            fn = row[COL_FN]; cod = row[COL_CODIGO]
            if pd.isna(fn): continue
            if fn > hoy:
                alertas.append({'Código':cod,'Centro':get_centro(cod),'Columna':COL_FN,
                                 'Valor':str(fn.date()),'Regla':'Fecha de nacimiento futura'})
                df.at[idx, COL_FN] = np.nan; continue
            edad = (hoy - fn).days / 365.25
            if edad < 10 or edad > 100:
                alertas.append({'Código':cod,'Centro':get_centro(cod),'Columna':COL_FN,
                                 'Valor':str(fn.date()),'Regla':f'Edad calculada = {edad:.1f} años (rango 10–100)'})
                df.at[idx, COL_FN] = np.nan

    cols_sem = [c for c in df.columns if '(0-7)' in c and 'Promedio' not in c]
    n_sem = 0
    for c in cols_sem:
        num = pd.to_numeric(df[c], errors='coerce')
        mask = num > 7
        for idx in df[mask].index:
            alertas.append({'Código':df.at[idx,COL_CODIGO],'Centro':get_centro(df.at[idx,COL_CODIGO]),
                             'Columna':c,'Valor':df.at[idx,c],'Regla':f'Días semanales > 7'})
            df.at[idx, c] = np.nan; n_sem += 1

    cols_mes = [c for c in df.columns if 'Total (0-28)' in c and 'Promedio' not in c]
    n_mes = 0
    for c in cols_mes:
        num = pd.to_numeric(df[c], errors='coerce')
        for idx in df[(num > 28) | (num < 0)].index:
            val = df.at[idx, c]
            alertas.append({'Código':df.at[idx,COL_CODIGO],'Centro':get_centro(df.at[idx,COL_CODIGO]),
                             'Columna':c,'Valor':val,'Regla':'Días mensuales fuera de 0–28'})
            df.at[idx, c] = np.nan; n_mes += 1

    logs.append(f"✓ Validación: {len(alertas)} valores corregidos "
                f"({n_sem} semanales, {n_mes} mensuales)")

    # ── Período ────────────────────────────────────────────────────────────────
    anio_actual = hoy.year
    fechas_validas = df[COL_FECHA].dropna()
    fechas_validas = fechas_validas[(fechas_validas.dt.year >= anio_actual-10) &
                                    (fechas_validas.dt.year <= anio_actual+1)]
    if len(fechas_validas):
        f_min, f_max = fechas_validas.min(), fechas_validas.max()
        if f_min.year == f_max.year and f_min.month == f_max.month:
            periodo = f'{MESES_ES[f_min.month]} {f_min.year}'
        elif f_min.year == f_max.year:
            periodo = f'{MESES_ES[f_min.month]}–{MESES_ES[f_max.month]} {f_min.year}'
        else:
            periodo = f'{MESES_ES[f_min.month]} {f_min.year} – {MESES_ES[f_max.month]} {f_max.year}'
    else:
        periodo = 'Período no determinado'
    logs.append(f"✓ Período detectado: {periodo}")

    # ── Construir Wide ─────────────────────────────────────────────────────────
    df = df.sort_values([COL_CODIGO, COL_FECHA]).reset_index(drop=True)
    conteo    = df[COL_CODIGO].value_counts()
    N_total   = int(conteo.shape[0])
    N_top2    = int((conteo >= 2).sum())
    N_solo1   = N_total - N_top2

    top1_rows, top2_rows = [], []
    for cod, grp in df.groupby(COL_CODIGO, sort=False):
        grp = grp.reset_index(drop=True)
        top1_rows.append(grp.loc[0])
        if len(grp) >= 2: top2_rows.append(grp.loc[1])

    df_top1 = pd.DataFrame(top1_rows).reset_index(drop=True)
    df_top2 = pd.DataFrame(top2_rows).reset_index(drop=True)
    df_top2_alin = (df_top2.set_index(COL_CODIGO)
                    .reindex(df_top1[COL_CODIGO]).reset_index())

    otras_cols = [c for c in df_top1.columns if c != COL_CODIGO]
    t1 = df_top1.rename(columns={c: f'{c}_TOP1' for c in otras_cols})
    t2 = df_top2_alin.rename(columns={c: f'{c}_TOP2' for c in otras_cols})
    wide = t1.merge(t2, on=COL_CODIGO, how='left')
    wide.insert(1, 'Tiene_TOP2',
        wide[[c for c in wide.columns if c.endswith('_TOP2')]]
        .notna().any(axis=1).map({True: 'Sí', False: 'No'}))
    logs.append(f"✓ Base wide: {len(wide)} filas × {len(wide.columns)} columnas")

    # ── Alertas seguimiento ────────────────────────────────────────────────────
    _HOY = pd.Timestamp.now().normalize()
    _col_fecha_top1 = next((c for c in wide.columns
                            if 'fecha entrevista' in c.lower() and c.endswith('_TOP1')), None)
    _n_rojo = _n_naranja = _n_verde = 0
    if _col_fecha_top1:
        _f1   = pd.to_datetime(wide[_col_fecha_top1], errors='coerce')
        _dias = (_HOY - _f1).dt.days
        def _alerta(d):
            if pd.isna(d): return ''
            if d < 60: return '🟢 <60 dias'
            if d < 90: return '🟠 60-89 dias'
            return '🔴 90+ dias'
        wide['Dias_desde_TOP1'] = _dias.where(wide['Tiene_TOP2'] == 'No', other=None)
        wide['Alerta_TOP2']     = _dias.where(wide['Tiene_TOP2'] == 'No').apply(
                                      lambda d: _alerta(d) if not pd.isna(d) else '')
        wide.loc[wide['Tiene_TOP2'] == 'Si', 'Alerta_TOP2'] = 'Completado'
        _n_rojo    = int((wide['Alerta_TOP2'] == '🔴 90+ dias').sum())
        _n_naranja = int((wide['Alerta_TOP2'] == '🟠 60-89 dias').sum())
        _n_verde   = int((wide['Alerta_TOP2'] == '🟢 <60 dias').sum())
    else:
        wide['Dias_desde_TOP1'] = None
        wide['Alerta_TOP2'] = ''

    # ── Normalizar sustancia ───────────────────────────────────────────────────
    for _sfx in ('_TOP1', '_TOP2'):
        _col_orig = None
        for c in wide.columns:
            if not c.endswith(_sfx): continue
            if any(k in _norm_str(c) for k in _SUST_KEYS_NORM):
                _col_orig = c; break
        if _col_orig is None: continue
        _col_raw = _col_orig.replace(_sfx, f'_RAW{_sfx}')
        wide.rename(columns={_col_orig: _col_raw}, inplace=True)
        wide[_col_orig] = wide[_col_raw].apply(norm_sust_v3)
        _idx = wide.columns.get_loc(_col_raw)
        _order = (list(wide.columns[:_idx+1]) + [_col_orig] +
                  [c for c in wide.columns[_idx+1:] if c != _col_orig])
        wide = wide[_order]
        logs.append(f"✓ Sustancia normalizada {_sfx}")

    # ── Duplicados ────────────────────────────────────────────────────────────
    dupes_data = []
    if _col_fecha_top1:
        raw_fecha = next((c for c in df.columns if _norm_str(c) in
                          [_norm_str(COL_FECHA)]), COL_FECHA)
        dup_mask = df.duplicated(subset=[COL_CODIGO, COL_FECHA], keep=False)
        dupes_df = df[dup_mask][[COL_CODIGO, COL_FECHA]].drop_duplicates()
        for _, row in dupes_df.iterrows():
            dupes_data.append({'Código': row[COL_CODIGO],
                                'Fecha': str(row[COL_FECHA])[:10]})

    N_dupes = len(dupes_data)
    logs.append(f"✓ Duplicados: {N_dupes} pacientes con fecha duplicada")

    # ── Generar Excel ─────────────────────────────────────────────────────────
    excel_bytes = _generar_excel(
        wide=wide, alertas=alertas, dupes=dupes_data,
        COL_CODIGO=COL_CODIGO, COL_CENTRO=COL_CENTRO,
        _col_fecha_top1=_col_fecha_top1,
        N_total=N_total, N_top2=N_top2, N_solo1=N_solo1,
        N_alertas=len(alertas), N_dupes=N_dupes,
        _n_rojo=_n_rojo, _n_naranja=_n_naranja, _n_verde=_n_verde,
        periodo=periodo)
    logs.append("✓ Excel generado con 6 hojas")

    # ── Estadísticas sustancia ─────────────────────────────────────────────────
    col_sust_top1 = next((c for c in wide.columns
                          if any(k in _norm_str(c) for k in _SUST_KEYS_NORM)
                          and c.endswith('_TOP1') and 'RAW' not in c), None)
    sust_dist = {}
    if col_sust_top1:
        sust_dist = (wide[col_sust_top1].dropna()
                     .value_counts().head(8).to_dict())

    # ── Centros — tabla resumen completa ──────────────────────────────────────
    centros = []
    if COL_CENTRO:
        col_centro_wide = f'{COL_CENTRO}_TOP1'
        if col_centro_wide in wide.columns:
            # Aplicaciones = filas en el df original por centro
            apps_por_centro = (df.groupby(COL_CENTRO)
                                 .size().reset_index(name='Aplicaciones')
                                 .rename(columns={COL_CENTRO: 'Centro'}))

            # Pacientes únicos, Con TOP2, Sin TOP2
            resumen = wide.groupby(col_centro_wide).agg(
                Pacientes=(COL_CODIGO, 'count'),
                Con_TOP2=('Tiene_TOP2', lambda x: (x=='Sí').sum())
            ).reset_index().rename(columns={col_centro_wide: 'Centro'})
            resumen['Sin_TOP2'] = resumen['Pacientes'] - resumen['Con_TOP2']

            # Valores corregidos por centro
            if alertas:
                import pandas as _pd2
                df_al = _pd2.DataFrame(alertas)
                corr_centro = (df_al.groupby('Centro').size()
                                    .reset_index(name='Vals_corregidos'))
                resumen = resumen.merge(corr_centro, on='Centro', how='left')
            else:
                resumen['Vals_corregidos'] = 0
            resumen['Vals_corregidos'] = resumen['Vals_corregidos'].fillna(0).astype(int)

            # Merge aplicaciones
            resumen = resumen.merge(apps_por_centro, on='Centro', how='left')
            resumen['Aplicaciones'] = resumen['Aplicaciones'].fillna(0).astype(int)

            # Ordenar por Aplicaciones desc
            resumen = resumen.sort_values('Aplicaciones', ascending=False)

            # Totales
            totales = {
                'Centro': 'TOTAL',
                'Aplicaciones': int(resumen['Aplicaciones'].sum()),
                'Pacientes': int(resumen['Pacientes'].sum()),
                'Con_TOP2': int(resumen['Con_TOP2'].sum()),
                'Sin_TOP2': int(resumen['Sin_TOP2'].sum()),
                'Vals_corregidos': int(resumen['Vals_corregidos'].sum()),
            }
            centros = resumen[['Centro','Aplicaciones','Pacientes',
                                'Con_TOP2','Sin_TOP2','Vals_corregidos']
                              ].to_dict('records')
            centros.append(totales)

    return {
        'wide':           wide,
        'filtro_centro':  filtro_centro,
        'fecha_desde':    fecha_desde,
        'fecha_hasta':    fecha_hasta,
        'stats': {
            'N_total':   N_total,
            'N_top2':    N_top2,
            'N_solo1':   N_solo1,
            'pct_top2':  round(N_top2/N_total*100,1) if N_total else 0,
            'N_alertas': len(alertas),
            'N_dupes':   N_dupes,
            'n_rojo':    _n_rojo,
            'n_naranja': _n_naranja,
            'n_verde':   _n_verde,
            'cols_wide': len(wide.columns),
            'sust_dist': sust_dist,
        },
        'centros':      centros,
        'alertas':      alertas,
        'dupes':        dupes_data,
        'periodo':      periodo,
        'excel_bytes':  excel_bytes,
        'logs':         logs,
    }


def _generar_excel(wide, alertas, dupes, COL_CODIGO, COL_CENTRO,
                   _col_fecha_top1, N_total, N_top2, N_solo1, N_alertas,
                   N_dupes, _n_rojo, _n_naranja, _n_verde, periodo):
    """Genera el Excel con 6 hojas y retorna BytesIO."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import pandas as pd

    _HOY = pd.Timestamp.now()
    thin = Side(style='thin', color=C_BDR)
    def bd(): return Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ── Hoja 1: Base Wide ─────────────────────────────────────────────────────
    ws = wb.active; ws.title = 'Base Wide'
    ws.freeze_panes = 'B3'
    ws.sheet_view.showGridLines = False

    # Fila 1: header
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f'A1:{_col_letter(len(wide.columns))}1')
    c = ws['A1']
    c.value = f'BASE WIDE TOP  ·  {N_total} pacientes  ·  {N_top2} con seguimiento  ·  {periodo}'
    c.font  = Font(bold=True, size=12, color=C_WHITE, name='Arial')
    c.fill  = PatternFill('solid', start_color=C_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # Fila 2: headers de columnas
    ws.row_dimensions[2].height = 22
    for ci, col in enumerate(wide.columns, 1):
        c = ws.cell(2, ci)
        c.value = col
        top2 = col.endswith('_TOP2') or col.endswith('_TOP1')
        c.font  = Font(bold=True, size=8, color=C_WHITE, name='Arial')
        c.fill  = PatternFill('solid', start_color=C_MID if '_TOP1' in col
                              else C_TOP2 if '_TOP2' in col else C_DARK)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bd()
        ws.column_dimensions[c.column_letter].width = 14

    # Datos (sin merge — directo a celdas individuales)
    for ri, (_, row) in enumerate(wide.iterrows(), 3):
        ws.row_dimensions[ri].height = 15
        for ci, val in enumerate(row, 1):
            try:
                c = ws.cell(ri, ci)
                if isinstance(val, float) and not pd.isna(val) and val == int(val):
                    c.value = int(val)
                elif not isinstance(val, str) and pd.isna(val):
                    c.value = None
                else:
                    c.value = val
                c.font      = Font(size=8, name='Arial')
                c.fill      = PatternFill('solid', start_color=C_ALT if ri%2==0 else C_WHITE)
                c.alignment = Alignment(vertical='center')
                c.border    = bd()
            except AttributeError:
                pass  # MergedCell — skip

    # ── Hoja 2: Resumen ───────────────────────────────────────────────────────
    wr = wb.create_sheet('Resumen'); wr.sheet_view.showGridLines = False
    wr.column_dimensions['A'].width = 3
    wr.column_dimensions['B'].width = 30
    wr.column_dimensions['C'].width = 18

    def _kpi(ws, row, label, value, color=C_DARK):
        ws.row_dimensions[row].height = 32
        ws.merge_cells(f'B{row}:C{row}')
        cl = ws[f'B{row}']; cl.value = label
        cl.font = Font(size=11, name='Arial', color='555555')
        cl.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        cv = ws[f'C{row}']; cv.value = value
        cv.font = Font(bold=True, size=16, name='Arial', color=color)
        cv.alignment = Alignment(horizontal='center', vertical='center')

    wr.merge_cells('B2:C2')
    c = wr['B2']
    c.value = f'RESUMEN  ·  {periodo}'
    c.font  = Font(bold=True, size=14, color=C_WHITE, name='Arial')
    c.fill  = PatternFill('solid', start_color=C_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    wr.row_dimensions[2].height = 32

    datos = [
        ('Pacientes únicos (ingreso)', N_total, C_DARK),
        ('Con seguimiento TOP2', f'{N_top2} ({round(N_top2/N_total*100,1) if N_total else 0}%)', C_MID),
        ('Solo TOP1 (pendientes)', N_solo1, '888888'),
        ('Período detectado', periodo, C_DARK),
        ('Valores corregidos (alertas)', N_alertas, 'C00000' if N_alertas else '538135'),
        ('Pacientes con fecha duplicada', N_dupes, 'C00000' if N_dupes else '538135'),
        ('🔴 Urgentes (90+ días sin TOP2)', _n_rojo, 'C00000'),
        ('🟠 Próximos (60–89 días)', _n_naranja, 'E67E22'),
        ('🟢 En plazo (<60 días)', _n_verde, '538135'),
    ]
    for i, (lbl, val, col) in enumerate(datos, 4):
        wr.row_dimensions[i].height = 24
        cl = wr[f'B{i}']; cl.value = lbl
        cl.font = Font(size=10, name='Arial', color='333333')
        cl.fill = PatternFill('solid', start_color=C_ALT if i%2==0 else C_WHITE)
        cl.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        cv = wr[f'C{i}']; cv.value = val
        cv.font = Font(bold=True, size=11, name='Arial', color=col)
        cv.fill = PatternFill('solid', start_color=C_ALT if i%2==0 else C_WHITE)
        cv.alignment = Alignment(horizontal='center', vertical='center')

    # ── Hoja 3: Alertas ───────────────────────────────────────────────────────
    wa = wb.create_sheet('Alertas'); wa.sheet_view.showGridLines = False
    wa.merge_cells('A1:E1')
    c = wa['A1']
    c.value = f'ALERTAS DE CALIDAD  ·  {N_alertas} valores corregidos a NaN'
    c.font = Font(bold=True, size=11, color=C_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color='C00000' if N_alertas else '538135')
    c.alignment = Alignment(horizontal='center', vertical='center')
    wa.row_dimensions[1].height = 24
    if alertas:
        df_al = pd.DataFrame(alertas)
        for ci, col in enumerate(['Código','Centro','Columna','Valor','Regla'], 1):
            c = wa.cell(2, ci); c.value = col
            c.font = Font(bold=True, size=9, color=C_WHITE, name='Arial')
            c.fill = PatternFill('solid', start_color=C_DARK)
            c.alignment = Alignment(horizontal='center', vertical='center')
        for ri, row in df_al.iterrows():
            for ci, col in enumerate(['Código','Centro','Columna','Valor','Regla'], 1):
                c = wa.cell(ri+3, ci); c.value = str(row.get(col,''))
                c.font = Font(size=8, name='Arial')
                c.fill = PatternFill('solid', start_color=C_ALT if ri%2==0 else C_WHITE)
        for col_letter, w in zip('ABCDE', [20,30,40,15,45]):
            wa.column_dimensions[col_letter].width = w
    else:
        wa['A3'].value = '✅ Sin alertas de calidad detectadas'

    # ── Hoja 4: Calidad de Datos ──────────────────────────────────────────────
    wq = wb.create_sheet('Calidad de Datos'); wq.sheet_view.showGridLines = False
    wq.merge_cells('A1:C1')
    c = wq['A1']
    c.value = f'CALIDAD DE DATOS  ·  {N_dupes} pacientes con fecha duplicada'
    c.font = Font(bold=True, size=11, color=C_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color='C00000' if N_dupes else '538135')
    c.alignment = Alignment(horizontal='center', vertical='center')
    wq.row_dimensions[1].height = 24
    if dupes:
        df_d = pd.DataFrame(dupes)
        for ci, col in enumerate(df_d.columns, 1):
            c = wq.cell(2, ci); c.value = col
            c.font = Font(bold=True, size=9, color=C_WHITE, name='Arial')
            c.fill = PatternFill('solid', start_color=C_DARK)
        for ri, row in df_d.iterrows():
            for ci, val in enumerate(row, 1):
                c = wq.cell(ri+3, ci); c.value = str(val)
                c.font = Font(size=8, name='Arial')
    else:
        wq['A3'].value = '✅ Sin fechas duplicadas detectadas'

    # ── Hoja 5: Por Centro ────────────────────────────────────────────────────
    wc = wb.create_sheet('Por Centro'); wc.sheet_view.showGridLines = False
    wc.merge_cells('A1:D1')
    c = wc['A1']
    c.value = 'RESUMEN POR CENTRO'
    c.font = Font(bold=True, size=11, color=C_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color=C_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    wc.row_dimensions[1].height = 24
    col_centro_wide = f'{COL_CENTRO}_TOP1' if COL_CENTRO else None
    if col_centro_wide and col_centro_wide in wide.columns:
        df_pc = wide.groupby(col_centro_wide).agg(
            Pacientes=(COL_CODIGO, 'count'),
            Con_TOP2=('Tiene_TOP2', lambda x: (x=='Sí').sum()),
        ).reset_index()
        df_pc['Sin_TOP2'] = df_pc['Pacientes'] - df_pc['Con_TOP2']
        df_pc.columns = ['Centro','Pacientes','Con TOP2','Sin TOP2']
        for ci, col in enumerate(df_pc.columns, 1):
            c = wc.cell(2, ci); c.value = col
            c.font = Font(bold=True, size=9, color=C_WHITE, name='Arial')
            c.fill = PatternFill('solid', start_color=C_MID)
        for ri, row in df_pc.iterrows():
            for ci, val in enumerate(row, 1):
                c = wc.cell(ri+3, ci); c.value = val
                c.font = Font(size=9, name='Arial')
                c.fill = PatternFill('solid', start_color=C_ALT if ri%2==0 else C_WHITE)
        for col_letter, w in zip('ABCD', [36,14,12,12]):
            wc.column_dimensions[col_letter].width = w
    else:
        wc['A3'].value = 'No se detectó columna de centro'

    # ── Hoja 6: Pendientes TOP2 ───────────────────────────────────────────────
    _pendientes = wide[wide['Alerta_TOP2'].isin(['🟠 60-89 dias','🔴 90+ dias'])].copy()
    wp2 = wb.create_sheet('Pendientes TOP2')
    wp2.sheet_view.showGridLines = False
    wp2.merge_cells('A1:E1')
    c = wp2['A1']
    c.value = f'PENDIENTES TOP2  ·  {len(_pendientes)} pacientes  ·  🔴 {_n_rojo}  🟠 {_n_naranja}'
    c.font = Font(bold=True, size=11, color=C_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color='C00000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    wp2.row_dimensions[1].height = 24
    if len(_pendientes):
        _cols_pend = [COL_CODIGO]
        if _col_fecha_top1: _cols_pend.append(_col_fecha_top1)
        _cols_pend += ['Dias_desde_TOP1','Alerta_TOP2']
        if col_centro_wide and col_centro_wide in _pendientes.columns:
            _cols_pend.insert(0, col_centro_wide)
        _tab = _pendientes[[c for c in _cols_pend if c in _pendientes.columns]].copy()
        _tab['_ord'] = _tab['Alerta_TOP2'].apply(lambda x: 0 if '90' in str(x) else 1)
        _tab = _tab.sort_values(['_ord','Dias_desde_TOP1'], ascending=[True,False]).drop(columns='_ord')
        for ci, col in enumerate(_tab.columns, 1):
            c = wp2.cell(2, ci); c.value = col
            c.font = Font(bold=True, size=9, color=C_WHITE, name='Arial')
            c.fill = PatternFill('solid', start_color='2F3640')
        for row_num, (_, row) in enumerate(_tab.iterrows(), 3):
            for ci, val in enumerate(row, 1):
                try:
                    c = wp2.cell(row_num, ci)
                    c.value = int(val) if isinstance(val, float) and not pd.isna(val) and val==int(val) else \
                              (None if not isinstance(val,str) and pd.isna(val) else val)
                    alerta_v = row.get('Alerta_TOP2','')
                    es_rojo = '90' in str(alerta_v)
                    c.font = Font(size=8, name='Arial',
                                  bold=(ci==len(_tab.columns)),
                                  color=('C00000' if es_rojo else 'E67E22') if 'Alerta' in str(_tab.columns[ci-1]) else '222222')
                    c.fill = PatternFill('solid', start_color='FDECEA' if es_rojo else 'FEF3E2')
                except (AttributeError, ValueError):
                    pass
        for col_letter, w in zip('ABCDE', [36,20,14,16,16]):
            wp2.column_dimensions[col_letter].width = w
    else:
        wp2['A3'].value = '✅ Sin pendientes urgentes'

    # ── Guardar a BytesIO ─────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _col_letter(n):
    """Convierte número de columna a letra (1→A, 27→AA)"""
    result = ''
    while n:
        n, rem = divmod(n-1, 26)
        result = chr(65+rem) + result
    return result
