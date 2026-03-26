"""
pipeline/runner.py — importa funciones directamente, sin subprocesos ni rutas en código.
"""
import sys, os, re, tempfile, shutil, importlib.util, types
from io import BytesIO
from pathlib import Path

PIPELINE_DIR = Path(__file__).parent

OUTPUTS = {
    'caract_excel': ('TOP_Caracterizacion_Ingreso.xlsx','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
    'seg_excel':    ('TOP_Seguimiento.xlsx','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
    'pdf_caract':   ('TOP_Informe_Caracterizacion.docx','application/vnd.openxmlformats-officedocument.wordprocessingml.document'),
    'pdf_seg':      ('TOP_Informe_Seguimiento.docx','application/vnd.openxmlformats-officedocument.wordprocessingml.document'),
    'pptx_caract':  ('TOP_Presentacion_Caracterizacion.pptx','application/vnd.openxmlformats-officedocument.presentationml.presentation'),
    'pptx_seg':     ('TOP_Presentacion_Seguimiento.pptx','application/vnd.openxmlformats-officedocument.presentationml.presentation'),
}

SCRIPT_FILES = {
    'caract_excel': 'caract_excel.py',
    'seg_excel':    'seg_excel.py',
    'pdf_caract':   'word_caract.py',
    'pdf_seg':      'word_seg.py',
    'pptx_caract':  'pptx_caract.py',
    'pptx_seg':     'pptx_seg.py',
}


def _load_mod(script_key, wide_path, out_path, filtro_centro=None):
    """
    Carga un script como módulo Python sin ejecutar código de nivel superior
    que dependa de rutas de archivos.
    """
    import types, re as _re
    src = open(str(PIPELINE_DIR / SCRIPT_FILES[script_key]), encoding='utf-8').read()

    # 1. Neutralizar INPUT_FILE = auto_archivo_wide()
    src = _re.sub(
        r'INPUT_FILE\s*=\s*auto_archivo_wide\(\)',
        'INPUT_FILE = None',
        src
    )
    # 2. Neutralizar OUTPUT_FILE hardcodeado
    for old in ["'/home/claude/TOP_Caracterizacion_Ingreso.xlsx'",
                "'/home/claude/TOP_Informe_Caracterizacion.pdf'",
                "'/home/claude/TOP_Informe_Seguimiento.pdf'",
                "'/home/claude/TOP_Presentacion_Caracterizacion.pptx'",
                "'/home/claude/TOP_Presentacion_Seguimiento.pptx'",
                "'/home/claude/TOP_Seguimiento.xlsx'"]:
        src = src.replace(old, 'None')

    # 3. Neutralizar llamadas a nivel de módulo que usen INPUT_FILE
    #    (como _detectar_pais(INPUT_FILE) o _pais_detectado = ...)
    src = _re.sub(
        r'^(_pais_detectado\s*=\s*_detectar_pais\(INPUT_FILE\))',
        r'_pais_detectado = None  # neutralizado',
        src, flags=_re.MULTILINE
    )

    # Crear módulo limpio y ejecutar
    mod = types.ModuleType('_qmod')
    mod.__file__ = str(PIPELINE_DIR / SCRIPT_FILES[script_key])
    try:
        exec(compile(src, '<qalat>', 'exec'), mod.__dict__)
    except SystemExit:
        pass
    except Exception as _e:
        import traceback
        print(f'[_load_mod] warning al cargar {script_key}: {_e}')
        traceback.print_exc()

    # Inyectar rutas y variables correctas después de cargar
    mod.__dict__['INPUT_FILE']       = wide_path
    mod.__dict__['OUTPUT_FILE']      = out_path
    mod.__dict__['FILTRO_CENTRO']    = filtro_centro
    mod.__dict__['NOMBRE_SERVICIO']  = 'Servicio de Tratamiento'
    mod.__dict__['PERIODO']          = ''
    mod.__dict__['auto_archivo_wide'] = lambda: wide_path

    return mod


def run_script(script_key, wide_path, filtro_centro=None):
    out_filename, mimetype = OUTPUTS[script_key]
    if filtro_centro:
        base, ext = out_filename.rsplit('.', 1)
        out_filename = f'{base}_{filtro_centro}.{ext}'

    # Archivo de salida temporal
    suffix = '.' + out_filename.rsplit('.', 1)[1]
    fd, out_path = tempfile.mkstemp(suffix=suffix, prefix='qalat_out_')
    os.close(fd)

    try:
        if script_key == 'caract_excel':
            from openpyxl import Workbook
            mod = _load_mod(script_key, wide_path, out_path)
            d, N = mod.cargar_ingreso()
            DC = mod.detectar_columnas(d.columns.tolist())
            wb = Workbook()
            mod.build_report(wb, d, N, DC)
            wb.save(out_path)

        elif script_key == 'seg_excel':
            from openpyxl import Workbook
            mod = _load_mod(script_key, wide_path, out_path)
            seg, N_total, N_seg, seg_tiempo = mod.cargar_datos()
            DC = mod.detectar_columnas(seg.columns.tolist())
            wb = Workbook()
            mod.build_seguimiento(wb, seg, N_total, N_seg, DC, seg_tiempo)
            mod.build_cambio_consumo(wb, seg, N_seg, DC)
            wb.save(out_path)

        elif script_key in ('pdf_caract', 'pdf_seg'):
            import subprocess
            src = open(str(PIPELINE_DIR / SCRIPT_FILES[script_key]), encoding='utf-8').read()

            # 1. Parchear INPUT_FILE = None → usar variable de entorno
            src = re.sub(
                r'INPUT_FILE\s*=\s*None.*?# runner inyecta la ruta real',
                'INPUT_FILE = __import__("os").environ["QALAT_WIDE"]',
                src
            )
            # 2. Parchear auto_archivo_wide para usar env var
            src = re.sub(
                r'def auto_archivo_wide\(\):.*?return [^\n]+\n',
                'def auto_archivo_wide():\n    return __import__("os").environ["QALAT_WIDE"]\n',
                src, flags=re.DOTALL
            )
            # 3. Parchear OUTPUT_FILE = None → usar env var
            src = src.replace(
                'OUTPUT_FILE   = None   # runner inyecta la ruta real',
                'OUTPUT_FILE = __import__("os").environ["QALAT_OUT"]'
            )
            src = src.replace(
                'OUTPUT_FILE = None   # runner inyecta la ruta real',
                'OUTPUT_FILE = __import__("os").environ["QALAT_OUT"]'
            )
            # 4. Parchear FILTRO_CENTRO = None → usar env var (puede estar vacío)
            src = src.replace(
                'FILTRO_CENTRO = None   # runner inyecta el filtro si aplica',
                'FILTRO_CENTRO = __import__("os").environ.get("QALAT_CENTRO") or None'
            )

            # Guardar script parcheado y ejecutar
            fd2, tmp_py = tempfile.mkstemp(suffix='.py', prefix='qs_word_')
            os.close(fd2)
            with open(tmp_py, 'w', encoding='utf-8') as f:
                f.write(src)

            env = os.environ.copy()
            env['QALAT_WIDE']   = wide_path
            env['QALAT_OUT']    = out_path
            env['QALAT_CENTRO'] = filtro_centro or ''

            try:
                r = subprocess.run(
                    [sys.executable, tmp_py],
                    capture_output=True, text=True,
                    timeout=180, env=env
                )
                if r.returncode != 0:
                    raise RuntimeError(r.stderr[-2000:] or r.stdout[-2000:])
            finally:
                try: os.unlink(tmp_py)
                except: pass

        elif script_key in ('pptx_caract', 'pptx_seg'):
            import subprocess

            fd2, tmp_py = tempfile.mkstemp(suffix='.py', prefix='qs_pptx_')
            os.close(fd2)
            import shutil
            shutil.copy(str(PIPELINE_DIR / SCRIPT_FILES[script_key]), tmp_py)

            env = os.environ.copy()
            env['QALAT_WIDE']   = wide_path
            env['QALAT_OUT']    = out_path
            env['QALAT_CENTRO'] = filtro_centro or ''

            try:
                r = subprocess.run(
                    [sys.executable, tmp_py],
                    capture_output=True, text=True,
                    timeout=180, env=env
                )
                if r.returncode != 0:
                    raise RuntimeError(r.stderr[-2000:] or r.stdout[-2000:])
            finally:
                try: os.unlink(tmp_py)
                except: pass

            if not os.path.exists(out_path) or os.path.getsize(out_path) == 0:
                raise FileNotFoundError('El script PPT no generó salida')

            with open(out_path, 'rb') as f:
                data = f.read()
            return BytesIO(data), out_filename, mimetype

        elif script_key in ('pptx_caract_OLD', 'pptx_seg_OLD'):
            import subprocess
            src = open(str(PIPELINE_DIR / SCRIPT_FILES[script_key]),
                       encoding='utf-8').read()

            # Calcular directorio temporal para json/js auxiliares
            tmp_dir = os.path.dirname(out_path).replace('\\','/')

            # 1. Parchear auto_archivo_wide
            src = re.sub(
                r'def auto_archivo_wide\(\):.*?return [^\n]+\n',
                'def auto_archivo_wide():\n    import os as _os\n    return _os.environ["QALAT_WIDE"]\n',
                src, flags=re.DOTALL
            )
            # 2. Parchear glob directo
            src = src.replace(
                "glob.glob('/home/claude/TOP_Base_Wide.xlsx')",
                '[__import__("os").environ["QALAT_WIDE"]]'
            )
            # 3. Parchear OUTPUT_FILE
            for old_path in [
                "'/home/claude/TOP_Presentacion_Caracterizacion.pptx'",
                "'/home/claude/TOP_Presentacion_Seguimiento.pptx'",
                "'/home/claude/TOP_Informe_Caracterizacion.pdf'",
                "'/home/claude/TOP_Informe_Seguimiento.pdf'",
                "'/home/claude/TOP_Seguimiento.xlsx'",
                "'/home/claude/TOP_Caracterizacion_Ingreso.xlsx'",
            ]:
                src = src.replace(old_path, '__import__("os").environ["QALAT_OUT"]')

            # 4. Parchear rutas de archivos auxiliares JSON y JS a directorio tmp
            for old_aux, new_aux in [
                ("'/home/claude/_top_car_data.json'",
                 f'"{tmp_dir}/_top_car_data.json"'),
                ("'/home/claude/_top_data.json'",
                 f'"{tmp_dir}/_top_data.json"'),
                ("'/home/claude/_top_car_builder.js'",
                 f'"{tmp_dir}/_top_car_builder.js"'),
                ("'/home/claude/_top_builder.js'",
                 f'"{tmp_dir}/_top_builder.js"'),
                # Rutas dentro del JS (readFileSync)
                ("fs.readFileSync('/home/claude/_top_car_data.json'",
                 f"fs.readFileSync('{tmp_dir}/_top_car_data.json'"),
                ("fs.readFileSync('/home/claude/_top_data.json'",
                 f"fs.readFileSync('{tmp_dir}/_top_data.json'"),
            ]:
                src = src.replace(old_aux, new_aux)

            # 5. Parchear OUTPUT_FILE en bloque JS
            src = src.replace(
                'OUTPUT_FILE + r"""',
                '__import__("os").environ["QALAT_OUT"] + r"""'
            )

            # Guardar script parcheado
            fd2, tmp_py = tempfile.mkstemp(suffix='.py', prefix='qs_')
            os.close(fd2)
            with open(tmp_py, 'w', encoding='utf-8') as f:
                f.write(src)

            env = os.environ.copy()
            env['QALAT_WIDE'] = wide_path
            env['QALAT_OUT']  = out_path

            try:
                r = subprocess.run(
                    [sys.executable, tmp_py],
                    capture_output=True, text=True,
                    timeout=180, env=env
                )
                if r.returncode != 0:
                    raise RuntimeError(r.stderr[-1000:] or r.stdout[-1000:])
            finally:
                try: os.unlink(tmp_py)
                except: pass

        if not os.path.exists(out_path) or os.path.getsize(out_path) == 0:
            raise FileNotFoundError('El script no generó salida')

        with open(out_path, 'rb') as f:
            data = f.read()
        return BytesIO(data), out_filename, mimetype

    finally:
        try: os.unlink(out_path)
        except: pass


def run_all(wide_path, progress_cb=None):
    results = {}
    keys = list(OUTPUTS.keys())
    for i, key in enumerate(keys):
        if progress_cb: progress_cb(i, len(keys), key)
        try:
            buf, fname, mime = run_script(key, wide_path)
            results[key] = {'ok': True, 'buf': buf, 'fname': fname, 'mime': mime}
        except Exception as e:
            results[key] = {'ok': False, 'error': str(e)}
    if progress_cb: progress_cb(len(keys), len(keys), 'listo')
    return results


# ══════════════════════════════════════════════════════════════════════════════
# DISTRIBUCIÓN POR CENTROS
# ══════════════════════════════════════════════════════════════════════════════
import zipfile, unicodedata as _ud

def _slug(s):
    """Convierte nombre de centro a nombre de carpeta seguro."""
    s = _ud.normalize('NFD', str(s)).encode('ascii', 'ignore').decode()
    s = re.sub(r'[^\w\s-]', '', s).strip()
    s = re.sub(r'[\s]+', '_', s)
    return s[:60]

def _detectar_centros(wide_path):
    """Lee la hoja 'Por Centro' del Wide y devuelve lista de centros (sin TOTAL)."""
    import pandas as pd
    try:
        df = pd.read_excel(wide_path, sheet_name='Por Centro', header=2)
        col = df.columns[0]
        centros = [str(v).strip() for v in df[col].dropna()
                   if str(v).strip().upper() != 'TOTAL' and str(v).strip() != '']
        return centros
    except Exception:
        # Fallback: leer hoja Base Wide y detectar columna de centro
        df = pd.read_excel(wide_path, sheet_name='Base Wide', header=1)
        def _n(s): return _ud.normalize('NFD', str(s).lower()).encode('ascii','ignore').decode()
        col_c = next((c for c in df.columns if any(k in _n(c) for k in
                      ['codigo del centro','centro de tratamiento','servicio de tratamiento'])
                      and 'trabajo' not in _n(c) and 'estudio' not in _n(c)), None)
        if col_c:
            return sorted(df[col_c].dropna().astype(str).str.strip().unique().tolist())
        return []

def _filtrar_wide_centro(wide_path, centro, out_path):
    """
    Genera un Excel Wide filtrado por centro.
    Copia todas las hojas del Wide original pero filtra
    la hoja 'Base Wide' al centro indicado.
    """
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb_orig = load_workbook(wide_path)
    wb_new  = load_workbook(wide_path)  # copia completa para mantener estilos

    # Filtrar hoja Base Wide
    ws = wb_new['Base Wide']
    # Leer con pandas para filtrar
    df = pd.read_excel(wide_path, sheet_name='Base Wide', header=1)
    def _n(s): return _ud.normalize('NFD', str(s).lower()).encode('ascii','ignore').decode()
    col_c = next((c for c in df.columns if any(k in _n(c) for k in
                  ['codigo del centro','centro de tratamiento','servicio de tratamiento'])
                  and 'trabajo' not in _n(c) and 'estudio' not in _n(c)), None)
    if col_c:
        df_f = df[df[col_c].astype(str).str.strip() == str(centro)].copy().reset_index(drop=True)
    else:
        df_f = df.copy()

    # Reescribir hoja Base Wide con datos filtrados (manteniendo fila de encabezado original)
    # Eliminar filas de datos (fila 3 en adelante) y reescribir
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for r_idx, row in enumerate(dataframe_to_rows(df_f, index=False, header=False), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb_new.save(out_path)


def run_paquetes_centros(wide_path, keys_sel=None, progress_cb=None, raw_input_path=None):
    """
    Genera un ZIP maestro con una carpeta por centro.
    Cada carpeta contiene:
      - Base Wide filtrada (Excel)
      - Los reportes seleccionados (Excel, Word, PPT)

    Parámetros:
      wide_path   : ruta al archivo Wide completo
      keys_sel    : lista de script_keys a generar (None = todos)
      progress_cb : callback(centro_actual, total_centros, paso_actual)

    Retorna:
      BytesIO con el ZIP maestro
    """
    if keys_sel is None:
        keys_sel = list(OUTPUTS.keys())

    centros = _detectar_centros(wide_path)
    if not centros:
        raise ValueError('No se detectaron centros en la base Wide.')

    n_centros = len(centros)
    zip_buf = BytesIO()

    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for i, centro in enumerate(centros):
            slug = _slug(centro)
            carpeta = f'{slug}/'

            if progress_cb:
                progress_cb(i, n_centros, centro)

            # 1. Base Wide filtrada por centro — regenerar TODAS las hojas
            wide_centro_path = None
            try:
                if raw_input_path and os.path.exists(raw_input_path):
                    # Regenerar Wide completa filtrada por centro (todas las hojas correctas)
                    from pipeline.wide_top import procesar_wide as _pw
                    res_c = _pw(raw_input_path, filtro_centro=centro)
                    fd_w, wide_centro_path = tempfile.mkstemp(suffix='.xlsx', prefix='qalat_wc_')
                    os.close(fd_w)
                    with open(wide_centro_path, 'wb') as f:
                        f.write(res_c['excel_bytes'].getvalue())
                else:
                    # Fallback: filtrar solo hoja Base Wide
                    fd_w, wide_centro_path = tempfile.mkstemp(suffix='.xlsx', prefix='qalat_wc_')
                    os.close(fd_w)
                    _filtrar_wide_centro(wide_path, centro, wide_centro_path)

                with open(wide_centro_path, 'rb') as f:
                    zf.writestr(f'{carpeta}BASE_Wide_{slug}.xlsx', f.read())
            except Exception as e:
                zf.writestr(f'{carpeta}ERROR_base_{slug}.txt', f'Error generando base: {e}')
                wide_centro_path = None

            # 2. Reportes — usar Wide filtrada si existe, sino fallback con filtro
            effective_wide = wide_centro_path if wide_centro_path and os.path.exists(wide_centro_path) else wide_path
            use_filtro = None if (wide_centro_path and os.path.exists(wide_centro_path)) else centro

            for key in keys_sel:
                out_fname, _ = OUTPUTS[key]
                base_name = out_fname.rsplit('.', 1)[0]
                ext       = out_fname.rsplit('.', 1)[1]
                archivo_zip = f'{carpeta}{base_name}_{slug}.{ext}'
                try:
                    buf, _, _ = run_script(key, effective_wide, filtro_centro=use_filtro)
                    zf.writestr(archivo_zip, buf.getvalue())
                except Exception as e:
                    zf.writestr(f'{carpeta}ERROR_{key}_{slug}.txt',
                                f'Error generando {out_fname}: {e}')

            # Limpiar Wide temporal del centro
            if wide_centro_path and wide_centro_path != wide_path:
                try: os.unlink(wide_centro_path)
                except: pass

    if progress_cb:
        progress_cb(n_centros, n_centros, 'listo')

    zip_buf.seek(0)
    return zip_buf
