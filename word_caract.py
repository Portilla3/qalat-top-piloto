"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   SCRIPT_TOP_Universal_Seguimiento_Excel.py                                ║
║   Genera tablas comparativas TOP1 (Ingreso) vs TOP2 (Seguimiento)          ║
║   Compatible con cualquier país que use el instrumento TOP                 ║
║   Versión Universal 1.0                                                    ║
╠══════════════════════════════════════════════════════════════════════════════╣
║                                                                             ║
║  CÓMO USAR LA PRÓXIMA VEZ:                                                 ║
║  1. Abre un chat nuevo con Claude                                           ║
║  2. Sube DOS archivos:                                                      ║
║       • Este script: SCRIPT_TOP_Universal_Seguimiento_Excel.py             ║
║       • La base en formato Wide (generada por SCRIPT_TOP_Universal_Wide)   ║
║  3. Escribe exactamente:                                                    ║
║     "Ejecuta el script universal Seguimiento Excel con esta base Wide"     ║
║                                                                             ║
║  TABLAS GENERADAS:                                                          ║
║  Hoja 1 – TOP Seguimiento (8 tablas):                                      ║
║    1.  Sustancia Principal (TOP1 vs TOP2)                                  ║
║    2.  Promedio Días de Consumo por Sustancia (TOP1 vs TOP2)               ║
║    3.  % Consumidores por Sustancia (TOP1 vs TOP2)                         ║
║    4.  Autopercepción del Estado de Salud 0–20 (TOP1 vs TOP2)             ║
║    5.  Condiciones de Vivienda (TOP1 vs TOP2)                              ║
║    6.  Días Trabajados y Estudiados (TOP1 vs TOP2)                         ║
║    7.  Transgresión a la Norma Social (TOP1 vs TOP2)                       ║
║    8.  Tipos de Transgresión (TOP1 vs TOP2)                                ║
║  Hoja 2 – Cambio en Consumo:                                               ║
║    Abstinencia / Disminuyó / Sin cambio / Empeoró por sustancia            ║
║                                                                             ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN — Claude ajusta estas líneas según el archivo recibido
# ══════════════════════════════════════════════════════════════════════════════
INPUT_FILE  = '/home/claude/TOP_Base_Wide.xlsx'   # ← Base Wide
SHEET_NAME  = 'Base Wide'                          # ← Nombre de hoja Wide
OUTPUT_FILE = '/home/claude/TOP_Seguimiento.xlsx'
# ── FILTRO OPCIONAL POR CENTRO ────────────────────────────────────────────────
# Dejar en None para procesar TODOS los centros.
# Poner el código exacto del centro para filtrar solo ese centro.
# Ejemplos:
#   FILTRO_CENTRO = None         ← todos los centros
#   FILTRO_CENTRO = "HCHN01"     ← solo ese centro
FILTRO_CENTRO = None
# ─────────────────────────────────────────────────────────────────────────────


COL_CODIGO = 'Código de identificación (2 primeras letras del primer nombre, 2 primeras letras del primer apellido, y día, mes y año de nacimiento)'

# ══════════════════════════════════════════════════════════════════════════════
import glob, os, unicodedata

def _norm(s):
    return unicodedata.normalize('NFD', str(s).lower()).encode('ascii','ignore').decode()

def auto_archivo_wide():
    """Encuentra automáticamente la base Wide subida al chat"""
    candidatos = (
        glob.glob('/mnt/user-data/uploads/*Wide*.xlsx') +
        glob.glob('/mnt/user-data/uploads/*wide*.xlsx') +
        glob.glob('/mnt/user-data/uploads/TOP_Base*.xlsx') +
        glob.glob('/home/claude/TOP_Base_Wide.xlsx'))
    if not candidatos:
        raise FileNotFoundError(
            "\n\u26a0  No se encontró la base Wide.\n"
            "   Sube el archivo TOP_Base_Wide.xlsx junto con este script.")
    print(f"  \u2192 Base Wide detectada: {os.path.basename(candidatos[0])}")
    return candidatos[0]

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import warnings
warnings.filterwarnings('ignore')

def _es_positivo(valor):
    s = str(valor).strip().lower()
    if s in ('sí', 'si'): return True
    if s in ('no', 'no aplica', 'nunca', 'nan', ''): return False
    n = pd.to_numeric(valor, errors='coerce')
    return not pd.isna(n) and n > 0

C_DARK='1F3864'; C_MID='2E75B6'; C_LIGHT='BDD7EE'; C_WHITE='FFFFFF'
C_ALT='EEF4FB';  C_NOTE='595959'; C_BDR='B8CCE4'

thin = Side(style='thin', color=C_BDR)
def bd(): return Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Helpers de formato ────────────────────────────────────────────────────────
def sec(ws, row, num, title):
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f'B{row}:G{row}')
    c = ws[f'B{row}']; c.value = f'{num}.  {title}'
    c.font = Font(bold=True, size=11, color=C_WHITE, name='Arial')
    c.fill = PatternFill('solid', start_color=C_MID)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    for col in range(2, 8): ws.cell(row, col).border = bd()
    return row + 1

def hdrs(ws, row, labels):
    ws.row_dimensions[row].height = 28
    for cl, lbl in zip(['B','C','D','E','F','G'], labels):
        c = ws[f'{cl}{row}']; c.value = lbl
        c.font = Font(bold=True, size=9, color=C_DARK, name='Arial')
        c.fill = PatternFill('solid', start_color=C_LIGHT)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bd()
    return row + 1

def drow(ws, row, vals, alt=False, ba=False):
    ws.row_dimensions[row].height = 16
    bg = C_ALT if alt else C_WHITE
    for i, (cl, val) in enumerate(zip(['B','C','D','E','F','G'], vals)):
        c = ws[f'{cl}{row}']
        c.value = round(val, 1) if isinstance(val, float) and not np.isnan(val) else val
        es_flecha = val in ['↑','↓']
        c.font = Font(size=9, name='Arial', bold=(ba or es_flecha),
                      color='008000' if val == '↑' else ('CC0000' if val == '↓' else (C_WHITE if ba else '000000')))
        c.fill = PatternFill('solid', start_color=C_DARK if ba else bg)
        c.alignment = Alignment(horizontal='left' if i==0 else 'center',
                                vertical='center', indent=1 if i==0 else 0)
        c.border = bd()
    return row + 1

def note(ws, row, text):
    ws.merge_cells(f'B{row}:G{row}')
    c = ws[f'B{row}']; c.value = text
    c.font = Font(size=8, color=C_NOTE, name='Arial', italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 13
    return row + 2

def cambio(v1, v2, mejor_si_sube=True):
    if pd.isna(v1) or pd.isna(v2): return '–'
    if abs(v1 - v2) < 0.01: return '–'
    subio = v2 > v1
    return '↑' if (subio == mejor_si_sube) else '↓'

# ── Normalización sustancia principal ─────────────────────────────────────────
def norm_sust(s):
    if pd.isna(s) or str(s).strip() == '0': return None
    s = str(s).strip().lower()
    if any(x in s for x in ['alcohol','cerveza','licor','aguard','alchol','bebida']): return 'Alcohol'
    if any(x in s for x in ['marihu','marjhu','marhuana','cannabis','cannbis','cannabin']): return 'Cannabis/Marihuana'
    if any(x in s for x in ['pasta base','pasta','papelillo']): return 'Pasta Base'
    if any(x in s for x in ['crack','cristal','piedra','paco']): return 'Crack/Cristal'
    if any(x in s for x in ['cocain','perico','coca ']): return 'Cocaína'
    if any(x in s for x in ['tabaco','cigarr','nicot']): return 'Tabaco/Nicotina'
    if any(x in s for x in ['inhalant','thiner','activo','resistol','cemento']): return 'Inhalantes'
    if any(x in s for x in ['sedant','benzod','tranqui','valium','clonaz']): return 'Sedantes'
    if any(x in s for x in ['opiod','heroina','morfin','fentanil']): return 'Opiáceos'
    if any(x in s for x in ['metanfet','anfetam']): return 'Metanfetamina'
    return 'Otras'

# ══════════════════════════════════════════════════════════════════════════════
# DETECCIÓN DINÁMICA DE COLUMNAS
# Trabaja sobre columnas _TOP1 y busca su par _TOP2
# ══════════════════════════════════════════════════════════════════════════════
def detectar_columnas(cols):
    col_set = set(cols)

    def par(c1):
        """Devuelve (col_TOP1, col_TOP2) si ambas existen, sino (col_TOP1, None)."""
        c2 = c1.replace('_TOP1', '_TOP2')
        return (c1, c2 if c2 in col_set else None)

    # 1. Sustancias: "1) Registrar... >> Total (0-28)_TOP1"
    sust_cols = []
    for c in cols:
        if c.endswith('_TOP1') and 'Total (0-28)' in c:
            base = c.replace('_TOP1', '')
            if base.startswith('1)'):
                partes = base.split('>>')
                if len(partes) >= 3:
                    nombre = partes[-2].strip().split('(')[0].strip()
                    c1, c2 = par(c)
                    sust_cols.append((nombre, c1, c2))
    print(f'  Sustancias: {[s[0] for s in sust_cols]}')

    # 2. Transgresión Sí/No
    tr_sn = []
    for c in cols:
        if c.endswith('_TOP1') and c.replace('_TOP1','').startswith('3)') and '>>' in c:
            nombre = c.replace('_TOP1','').split('>>')[-1].strip()
            c1, c2 = par(c)
            tr_sn.append((nombre, c1, c2))
    print(f'  Transgresión: {[t[0] for t in tr_sn]}')

    def detectar(prefijo, contiene=None, termina=None):
        for c in cols:
            base = c.replace('_TOP1','')
            if not c.endswith('_TOP1'): continue
            if not base.startswith(prefijo): continue
            if contiene and contiene not in c.lower(): continue
            if termina and not c.replace('_TOP1','').endswith(termina): continue
            return par(c)
        return (None, None)

    vif      = detectar('4)', contiene='violencia intrafamiliar', termina='Total (0-28)')
    sal_psi  = detectar('6)')
    sal_fis  = detectar('8)')
    cal_vid  = detectar('10)')
    viv1     = detectar('9)', contiene='estable')
    viv2     = detectar('9)', contiene='básicas')
    trab     = detectar('7)', contiene='trabajo remunerado', termina='Total (0-28)')
    estud    = detectar('7)', contiene='colegio', termina='Total (0-28)') or \
               detectar('7)', contiene='instituto', termina='Total (0-28)')
    sust_ppal = next(
        (par(c) for c in cols if c.endswith('_TOP1')
         and c.replace('_TOP1','').startswith('2)')
         and 'sustancia principal' in c.lower()), (None, None))

    DC = dict(
        sust_cols=sust_cols, tr_sn=tr_sn,
        vif=vif, sal_psi=sal_psi, sal_fis=sal_fis, cal_vid=cal_vid,
        viv1=viv1, viv2=viv2, trab=trab, estud=estud, sust_ppal=sust_ppal
    )
    for k, v in DC.items():
        if isinstance(v, tuple) and v[0] is None and k not in ('sust_cols','tr_sn'):
            print(f'  ⚠️  No encontrada: {k}')
    return DC

# ══════════════════════════════════════════════════════════════════════════════
# CARGA — filtra solo pacientes con TOP2
# ══════════════════════════════════════════════════════════════════════════════
def cargar_datos():
    print(f'  Leyendo: {INPUT_FILE}  |  Hoja: {SHEET_NAME}')
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=1)

    # Aplicar filtro de centro si corresponde
    _col_centro = next((c for c in df.columns if any(x in _norm(c) for x in
                        ['codigo del centro', 'servicio de tratamiento',
                         'centro/ servicio', 'codigo centro'])), None)
    if FILTRO_CENTRO and _col_centro:
        n_antes = len(df)
        df = df[df[_col_centro].astype(str).str.strip() == FILTRO_CENTRO].copy()
        df = df.reset_index(drop=True)
        print(f'  ⚑ Filtro activo: Centro = "{FILTRO_CENTRO}"')
        print(f'    {n_antes} pacientes totales → {len(df)} del centro seleccionado')
    if FILTRO_CENTRO:
        global OUTPUT_FILE
        OUTPUT_FILE = f'/home/claude/TOP_Seguimiento_{FILTRO_CENTRO}.xlsx'
    N_total = len(df)
    seg = df[df['Tiene_TOP2'] == 'Sí'].copy().reset_index(drop=True)
    N_seg = len(seg)
    print(f'  Total pacientes:      {N_total}')
    print(f'  Con TOP2 (seguimiento): {N_seg}  ({round(N_seg/N_total*100,1)}%)')

    # ── Tiempo de seguimiento ────────────────────────────────────────────────
    _fc1 = next((c for c in seg.columns if 'fecha entrevista' in c.lower() and c.endswith('_TOP1')), None)
    _fc2 = next((c for c in seg.columns if 'fecha entrevista' in c.lower() and c.endswith('_TOP2')), None)
    seg_tiempo = {'mediana': None, 'min': None, 'max': None, 'n': 0, 'n_total': N_seg}
    if _fc1 and _fc2:
        _d1 = pd.to_datetime(seg[_fc1], errors='coerce')
        _d2 = pd.to_datetime(seg[_fc2], errors='coerce')
        _dias = (_d2 - _d1).dt.days
        _dias_ok = _dias[(_dias >= 0) & (_dias <= 730)].dropna()
        if len(_dias_ok) > 0:
            _m = _dias_ok / 30.44
            seg_tiempo = {
                'mediana': round(float(_m.median()), 1),
                'min':     round(float(_m.min()), 1),
                'max':     round(float(_m.max()), 1),
                'n':       len(_dias_ok),
                'n_total': int(_dias.notna().sum())
            }

    return seg, N_total, N_seg, seg_tiempo

# Helpers para leer columna con sufijo correcto
def v1(df, col): return pd.to_numeric(df[col], errors='coerce') if col else pd.Series(dtype=float)
def v2(df, col): return pd.to_numeric(df[col], errors='coerce') if col else pd.Series(dtype=float)
def si_sn(df, col): return df[col] == 'Sí' if col and col in df.columns else pd.Series([False]*len(df))

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1: TABLAS DE SEGUIMIENTO
# ══════════════════════════════════════════════════════════════════════════════
def build_seguimiento(wb, seg, N_total, N_seg, DC, seg_tiempo=None):
    ws = wb.active; ws.title = 'TOP - Seguimiento'
    ws.sheet_properties.tabColor = C_MID
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 44
    for col in ['C','D','E','F']: ws.column_dimensions[col].width = 13
    ws.column_dimensions['G'].width = 10

    # Tiempo de seguimiento — texto para encabezado
    if seg_tiempo and seg_tiempo['mediana'] is not None:
        n_v = seg_tiempo['n']; n_t = seg_tiempo['n_total']
        txt_tiempo = (f"Tiempo entre TOP1 y TOP2  ·  Mediana: {seg_tiempo['mediana']} meses  "
                      f"·  Rango: {seg_tiempo['min']}–{seg_tiempo['max']} meses  "
                      f"·  N válido: {n_v}" +
                      (f" de {n_t}" if n_t != n_v else ""))
    else:
        txt_tiempo = "Tiempo entre TOP1 y TOP2: no disponible"

    # Encabezado
    for r, txt, bg, sz, bold, tc in [
        (1, 'INFORME DE SEGUIMIENTO  ·  TOP',                                       C_DARK, 16, True,  C_WHITE),
        (2, f'Comparación TOP 1 (Ingreso) vs TOP 2 (Seguimiento)  ·  {N_seg} pacientes con ambas evaluaciones', C_MID, 10, False, C_WHITE),
        (3, f'Total pacientes en base: {N_total}  ·  Con seguimiento: {N_seg} ({round(N_seg/N_total*100,1)}%)  ·  % calculados sobre N válido', C_LIGHT, 9, False, C_DARK),
        (4, txt_tiempo, 'E2EFDA', 9, False, '375623'),
    ]:
        ws.merge_cells(f'B{r}:G{r}')
        c = ws[f'B{r}']; c.value = txt
        c.font = Font(bold=bold, size=sz, color=tc, name='Arial')
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 36 if r==1 else 20 if r==2 else 14

    R = 5

    # ── 1. SUSTANCIA PRINCIPAL ────────────────────────────────────────────────
    R = sec(ws, R, '1', 'Sustancia Principal de Problema')
    R = hdrs(ws, R, ['Sustancia', 'TOP 1\nn', 'TOP 1\n%', 'TOP 2\nn', 'TOP 2\n%', 'N válido'])
    c1_sp, c2_sp = DC['sust_ppal']
    if c1_sp:
        sr1 = seg[c1_sp].apply(norm_sust)
        sr2 = seg[c2_sp].apply(norm_sust) if c2_sp else pd.Series([None]*N_seg)
        nv1 = int(sr1.notna().sum()); nv2 = int(sr2.notna().sum())
        cats = ['Alcohol','Cannabis/Marihuana','Pasta Base','Cocaína','Crack/Cristal',
                'Tabaco/Nicotina','Inhalantes','Sedantes','Opiáceos','Metanfetamina','Otras']
        for i, cat in enumerate(cats):
            n1 = int((sr1==cat).sum()); n2 = int((sr2==cat).sum())
            if n1==0 and n2==0: continue
            R = drow(ws, R, [cat,
                n1, round(n1/nv1*100,1) if nv1>0 else 0,
                n2, round(n2/nv2*100,1) if nv2>0 else 0,
                f'{nv1} / {nv2}'], alt=i%2==0)
        R = drow(ws, R, ['N válido', nv1, '100%', nv2, '100%', ''], ba=True)
        R = note(ws, R, f'% sobre casos con sustancia identificada. TOP1 n={nv1}, TOP2 n={nv2}.')
    else:
        R = note(ws, R, 'Columna sustancia principal no encontrada.')

    # ── 2. DÍAS DE CONSUMO POR SUSTANCIA ─────────────────────────────────────
    R = sec(ws, R, '2', 'Promedio de Días de Consumo por Sustancia (últimas 4 semanas, 0–28)')
    R = hdrs(ws, R, ['Sustancia', 'TOP 1\nProm. días', 'N válido\nTOP 1', 'TOP 2\nProm. días', 'N válido\nTOP 2', 'Cambio'])
    for i, (lbl, c1, c2) in enumerate(DC['sust_cols']):
        s1 = v1(seg, c1); s2 = v2(seg, c2)
        m1 = float(s1.mean()) if s1.notna().sum()>0 else np.nan
        m2 = float(s2.mean()) if (c2 and s2.notna().sum()>0) else np.nan
        nv1_ = int(s1.notna().sum()); nv2_ = int(s2.notna().sum()) if c2 else 0
        ch = cambio(m1, m2, mejor_si_sube=False)
        R = drow(ws, R, [lbl,
            round(m1,1) if not np.isnan(m1) else 0, nv1_,
            round(m2,1) if not np.isnan(m2) else 0, nv2_, ch], alt=i%2==0)
    R = note(ws, R, f'Promedio sobre los {N_seg} pacientes con seguimiento (incluye 0). ↑ Reducción del consumo (mejora).')

    # ── 3. % CONSUMIDORES POR SUSTANCIA ──────────────────────────────────────
    R = sec(ws, R, '3', f'% de Personas que Consume cada Sustancia (sobre N={N_seg})')
    R = hdrs(ws, R, ['Sustancia', 'TOP 1\nn', 'TOP 1\n%', 'TOP 2\nn', 'TOP 2\n%', 'Cambio'])
    for i, (lbl, c1, c2) in enumerate(DC['sust_cols']):
        s1 = v1(seg, c1); s2 = v2(seg, c2)
        n1 = int((s1>0).sum()); n2 = int((s2>0).sum()) if c2 else 0
        p1 = round(n1/N_seg*100,1); p2 = round(n2/N_seg*100,1) if c2 else 0
        ch = cambio(p1, p2, mejor_si_sube=False)
        R = drow(ws, R, [lbl, n1, p1, n2, p2, ch], alt=i%2==0)
    R = note(ws, R, f'% sobre N={N_seg} pacientes. Puede sumar >100% (una persona puede consumir varias sustancias).')

    # ── 4. SALUD Y CALIDAD DE VIDA ────────────────────────────────────────────
    R = sec(ws, R, '4', 'Autopercepción del Estado de Salud y Calidad de Vida (escala 0–20)')
    R = hdrs(ws, R, ['Dimensión', 'TOP 1\nPromedio', 'N válido\nTOP 1', 'TOP 2\nPromedio', 'N válido\nTOP 2', 'Cambio'])
    for i, (lbl, (c1, c2)) in enumerate([
        ('Salud Psicológica', DC['sal_psi']),
        ('Salud Física',      DC['sal_fis']),
        ('Calidad de Vida',   DC['cal_vid']),
    ]):
        if c1 is None: continue
        s1 = v1(seg, c1); s2 = v2(seg, c2)
        m1 = float(s1.mean()); m2 = float(s2.mean()) if c2 else np.nan
        nv1_ = int(s1.notna().sum()); nv2_ = int(s2.notna().sum()) if c2 else 0
        ch = cambio(m1, m2, mejor_si_sube=True)
        R = drow(ws, R, [lbl, round(m1,1), nv1_, round(m2,1) if not np.isnan(m2) else 0, nv2_, ch], alt=i%2==0)
    R = note(ws, R, 'Escala 0 = Muy mal / 20 = Excelente. ↑ Mejora · ↓ Empeora.')

    # ── 5. VIVIENDA ───────────────────────────────────────────────────────────
    R = sec(ws, R, '5', 'Condiciones de Vivienda (últimas 4 semanas)')
    R = hdrs(ws, R, ['Condición', 'TOP 1\nn Sí', 'TOP 1\n% Sí', 'TOP 2\nn Sí', 'TOP 2\n% Sí', 'Cambio'])
    for i, (lbl, (c1, c2)) in enumerate([
        ('Lugar estable para vivir',        DC['viv1']),
        ('Vivienda con condiciones básicas', DC['viv2']),
    ]):
        if c1 is None: continue
        nv1_ = int(seg[c1].isin(['Sí','No']).sum()) or N_seg
        nv2_ = int(seg[c2].isin(['Sí','No']).sum()) if c2 else N_seg
        n1 = int((seg[c1]=='Sí').sum())
        n2 = int((seg[c2]=='Sí').sum()) if c2 else 0
        p1 = round(n1/nv1_*100,1); p2 = round(n2/nv2_*100,1) if c2 else 0
        ch = cambio(p1, p2, mejor_si_sube=True)
        R = drow(ws, R, [lbl, n1, p1, n2, p2, ch], alt=i%2==0)
    R = note(ws, R, '% sobre casos con respuesta Sí/No válida. ↑ Mejora.')

    # ── 6. DÍAS TRABAJADOS Y ESTUDIADOS ──────────────────────────────────────
    R = sec(ws, R, '6', 'Días Trabajados y Estudiados (últimas 4 semanas, 0–28)')
    R = hdrs(ws, R, ['Actividad', 'TOP 1\nProm. días', 'N válido\nTOP 1', 'TOP 2\nProm. días', 'N válido\nTOP 2', 'Cambio'])
    for i, (lbl, (c1, c2)) in enumerate([
        ('Días de trabajo remunerado',          DC['trab']),
        ('Días asistidos a inst. educativa',    DC['estud']),
    ]):
        if c1 is None: continue
        s1 = v1(seg, c1); s2 = v2(seg, c2)
        m1 = float(s1.mean()); m2 = float(s2.mean()) if c2 else np.nan
        nv1_ = int(s1.notna().sum()); nv2_ = int(s2.notna().sum()) if c2 else 0
        ch = cambio(m1, m2, mejor_si_sube=True)
        R = drow(ws, R, [lbl, round(m1,1), nv1_, round(m2,1) if not np.isnan(m2) else 0, nv2_, ch], alt=i%2==0)
    R = note(ws, R, 'Promedio sobre todos los pacientes (incluye 0). ↑ Mejora (más días de actividad).')

    # ── 7. TRANSGRESIÓN GENERAL ───────────────────────────────────────────────
    R = sec(ws, R, '7', 'Transgresión a la Norma Social (presencia de algún incidente)')
    R = hdrs(ws, R, ['Situación', 'TOP 1\nn', 'TOP 1\n%', 'TOP 2\nn', 'TOP 2\n%', 'N válido'])

    tr_cols1 = [c1 for _, c1, _ in DC['tr_sn']]
    tr_cols2 = [c2 for _, _, c2 in DC['tr_sn']]
    vif_c1, vif_c2 = DC['vif']

    def has_tr(row, sn_cols, vif_col):
        for c in sn_cols:
            if c and _es_positivo(row.get(c,'')): return True
        if vif_col:
            v = pd.to_numeric(row.get(vif_col, np.nan), errors='coerce')
            return not np.isnan(v) and v > 0
        return False

    tr1 = seg.apply(lambda r: int(has_tr(r, tr_cols1, vif_c1)), axis=1)
    tr2 = seg.apply(lambda r: int(has_tr(r, tr_cols2, vif_c2)), axis=1)
    n1_si = int(tr1.sum()); n2_si = int(tr2.sum())
    n1_no = N_seg - n1_si; n2_no = N_seg - n2_si

    R = drow(ws, R, ['Sí, cometió alguna transgresión',
        n1_si, round(n1_si/N_seg*100,1),
        n2_si, round(n2_si/N_seg*100,1),
        f'{N_seg} / {N_seg}'], alt=False)
    R = drow(ws, R, ['No, ninguna transgresión',
        n1_no, round(n1_no/N_seg*100,1),
        n2_no, round(n2_no/N_seg*100,1),
        f'{N_seg} / {N_seg}'], alt=True)
    R = drow(ws, R, ['N válido', N_seg, '100%', N_seg, '100%', ''], ba=True)
    tipos_str = ', '.join(n for n,_,_ in DC['tr_sn'])
    R = note(ws, R, f'Incluye {tipos_str} y VIF. ↑ Reducción de transgresión (mejora).')

    # ── 8. TIPOS DE TRANSGRESIÓN ──────────────────────────────────────────────
    R = sec(ws, R, '8', f'Tipos de Transgresión (% sobre N={N_seg} pacientes con seguimiento)')
    R = hdrs(ws, R, ['Tipo de transgresión', 'TOP 1\nn', 'TOP 1\n%', 'TOP 2\nn', 'TOP 2\n%', 'Cambio'])
    for i, (lbl, c1, c2) in enumerate(DC['tr_sn']):
        n1 = int(seg[c1].apply(_es_positivo).sum()) if c1 else 0
        n2 = int(seg[c2].apply(_es_positivo).sum()) if c2 else 0
        p1 = round(n1/N_seg*100,1); p2 = round(n2/N_seg*100,1)
        ch = cambio(p1, p2, mejor_si_sube=False)
        R = drow(ws, R, [lbl, n1, p1, n2, p2, ch], alt=i%2==0)
    if vif_c1:
        vif1_v = pd.to_numeric(seg[vif_c1], errors='coerce')
        vif2_v = pd.to_numeric(seg[vif_c2], errors='coerce') if vif_c2 else pd.Series([np.nan]*N_seg)
        n1v = int((vif1_v>0).sum()); n2v = int((vif2_v>0).sum())
        p1v = round(n1v/N_seg*100,1); p2v = round(n2v/N_seg*100,1)
        R = drow(ws, R, ['Violencia Intrafamiliar (VIF)', n1v, p1v, n2v, p2v,
                         cambio(p1v, p2v, mejor_si_sube=False)], alt=len(DC['tr_sn'])%2==0)
    R = note(ws, R, f'% sobre N={N_seg}. % no suman 100% (un paciente puede cometer más de un tipo). ↑ Reducción (mejora).')

    ws.freeze_panes = 'B5'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True; ws.page_setup.fitToWidth = 1
    print(f'  ✓ Hoja 1: 8 tablas  ·  N seguimiento = {N_seg}')

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2: CAMBIO EN CONSUMO
# ══════════════════════════════════════════════════════════════════════════════
def build_cambio_consumo(wb, seg, N_seg, DC):
    ws = wb.create_sheet('Cambio en Consumo')
    ws.sheet_properties.tabColor = C_DARK
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 22
    for col in ['C','D','E','F','G','H','I','J','K','L']:
        ws.column_dimensions[col].width = 10

    # Encabezado
    for r, txt, bg, sz, bold, tc in [
        (1, 'CAMBIO EN EL CONSUMO POR SUSTANCIA  ·  TOP  ·  Ingreso → Seguimiento', C_DARK, 13, True, C_WHITE),
        (2, f'Solo pacientes con consumo > 0 en TOP 1  ·  % sobre n consumidores  ·  N seguimiento = {N_seg}', C_MID, 9, False, C_WHITE),
    ]:
        ws.merge_cells(f'B{r}:L{r}')
        c = ws[f'B{r}']; c.value = txt
        c.font = Font(bold=bold, size=sz, color=tc, name='Arial')
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 28 if r==1 else 18

    # Cabeceras
    R = 4
    ws.row_dimensions[R].height = 28
    for cl, txt in zip(['B','C','D','E','F','G','H','I','J','K','L'],
                       ['Sustancia','n cons.\nTOP 1','Abstinencia\nn','Abstinencia\n%',
                        'Disminuyó\nn','Disminuyó\n%','Sin cambio\nn','Sin cambio\n%',
                        'Empeoró\nn','Empeoró\n%','% Abs +\nDisminuyó']):
        c = ws[f'{cl}{R}']; c.value = txt
        c.font = Font(bold=True, size=9, color=C_DARK, name='Arial')
        c.fill = PatternFill('solid', start_color=C_LIGHT)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bd()
    R += 1

    tot = {'n':0,'abs':0,'dis':0,'sc':0,'emp':0}
    for i, (lbl, c1, c2) in enumerate(DC['sust_cols']):
        if c2 is None: continue
        s1 = pd.to_numeric(seg[c1], errors='coerce').fillna(0)
        s2 = pd.to_numeric(seg[c2], errors='coerce').fillna(0)
        mask = s1 > 0
        n_cons = int(mask.sum())
        if n_cons < 2: continue
        sub1 = s1[mask]; sub2 = s2[mask]
        n_abs = int((sub2==0).sum())
        n_dis = int(((sub2>0)&(sub2<sub1)).sum())
        n_sc  = int((sub2==sub1).sum())
        n_emp = int((sub2>sub1).sum())
        pct   = lambda n: round(n/n_cons*100,1) if n_cons>0 else 0
        pabs_dis = round((n_abs+n_dis)/n_cons*100,1)

        bg = C_ALT if i%2==0 else C_WHITE
        ws.row_dimensions[R].height = 16
        for cl, val in zip(['B','C','D','E','F','G','H','I','J','K','L'],
                           [lbl, n_cons, n_abs, pct(n_abs), n_dis, pct(n_dis),
                            n_sc, pct(n_sc), n_emp, pct(n_emp), pabs_dis]):
            c = ws[f'{cl}{R}']; c.value = round(val,1) if isinstance(val,float) else val
            c.font = Font(size=9, name='Arial', bold=(cl=='L'),
                          color=C_MID if cl=='L' else '000000')
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='left' if cl=='B' else 'center',
                                    vertical='center', indent=1 if cl=='B' else 0)
            c.border = bd()
        tot['n']+=n_cons; tot['abs']+=n_abs; tot['dis']+=n_dis
        tot['sc']+=n_sc; tot['emp']+=n_emp
        R += 1

    # Fila de totales
    ws.row_dimensions[R].height = 18
    n_t = tot['n']; pct = lambda n: round(n/n_t*100,1) if n_t>0 else 0
    for cl, val in zip(['B','C','D','E','F','G','H','I','J','K','L'],
                       ['TOTAL (todas las sustancias)', n_t,
                        tot['abs'], pct(tot['abs']),
                        tot['dis'], pct(tot['dis']),
                        tot['sc'],  pct(tot['sc']),
                        tot['emp'], pct(tot['emp']),
                        round((tot['abs']+tot['dis'])/n_t*100,1) if n_t>0 else 0]):
        c = ws[f'{cl}{R}']; c.value = round(val,1) if isinstance(val,float) else val
        c.font = Font(bold=True, size=9, name='Arial', color=C_WHITE)
        c.fill = PatternFill('solid', start_color=C_DARK)
        c.alignment = Alignment(horizontal='left' if cl=='B' else 'center',
                                vertical='center', indent=1 if cl=='B' else 0)
        c.border = bd()
    R += 2

    ws.merge_cells(f'B{R}:L{R}')
    c = ws[f'B{R}']
    c.value = ('Abstinencia = consumo 0 en TOP2 (con consumo > 0 en TOP1). '
               'Disminuyó = TOP2 < TOP1. Sin cambio = TOP2 = TOP1. Empeoró = TOP2 > TOP1. '
               '% sobre n consumidores de esa sustancia en TOP1.')
    c.font = Font(size=8, color=C_NOTE, name='Arial', italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws.row_dimensions[R].height = 20

    ws.freeze_panes = 'B5'
    print(f'  ✓ Hoja 2: Cambio en Consumo')

# ══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    print('=' * 60)
    print('  SCRIPT_TOP_Universal_Seguimiento_Excel  —  Iniciando...')
    print('=' * 60)

    seg, N_total, N_seg, seg_tiempo = cargar_datos()

    print('\n→ Detectando columnas dinámicamente...')
    DC = detectar_columnas(seg.columns.tolist())

    print('\n→ Construyendo Excel...')
    wb = Workbook()
    build_seguimiento(wb, seg, N_total, N_seg, DC, seg_tiempo)
    build_cambio_consumo(wb, seg, N_seg, DC)
    wb.save(OUTPUT_FILE)

    print(f'\n{"=" * 60}')
    print(f'  ✅  LISTO  →  {OUTPUT_FILE}')
    print(f'{"=" * 60}')
